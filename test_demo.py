#!/usr/bin/env python3
"""
一键测试 Prajna Guangzhou Demo 的两个核心技能：
1. 薪资模板（prajna-salary-template）
2. 销售周报（prajna-sales-weekly-report）

运行方式：
    source venv/bin/activate
    python test_demo.py
"""

import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

APP_DIR = Path(__file__).resolve().parent


def run(cmd, **kwargs):
    """Run a command and print it."""
    print(f"\n>>> {' '.join(str(c) for c in cmd)}")
    result = subprocess.run(cmd, capture_output=True, text=True, **kwargs)
    if result.stdout:
        print(result.stdout.strip())
    if result.stderr:
        print(result.stderr.strip(), file=sys.stderr)
    result.check_returncode()
    return result


def test_salary():
    out = "/tmp/test_salary.xlsx"
    run(
        [
            sys.executable,
            str(APP_DIR / "skills" / "prajna-salary-template" / "scripts" / "generate_salary_template.py"),
            "--preset=互联网-电商运营助理",
            "--city=广州",
            "--level=P2",
            "--output", out,
        ]
    )
    wb = load_workbook(out)
    expected = [
        "薪资结构",
        "绩效考核",
        "调薪机制",
        "奖金方案",
        "福利明细",
        "计算示例",
        "合规校验",
        "绩效系数对照表",
    ]
    missing = [s for s in expected if s not in wb.sheetnames]
    assert not missing, f"薪资模板缺少工作表：{missing}"
    print(f"✅ 薪资模板测试通过：{out}")


def test_sales_weekly():
    out = "/tmp/test_sales_weekly.xlsx"
    run(
        [
            sys.executable,
            str(APP_DIR / "skills" / "prajna-sales-weekly-report" / "scripts" / "generate_sales_weekly_report.py"),
            "--preset=互联网/SaaS 销售团队",
            "--team=华南销售一部",
            "--week=2026年第30周",
            "--sales-target=1200000",
            "--author=销售主管",
            "--date=2026-07-20",
            "--output", out,
        ]
    )
    wb = load_workbook(out)
    expected = [
        "封面汇总",
        "核心业绩数据",
        "重点商机进展",
        "问题分析",
        "下周工作计划",
    ]
    missing = [s for s in expected if s not in wb.sheetnames]
    assert not missing, f"销售周报缺少工作表：{missing}"
    print(f"✅ 销售周报测试通过：{out}")


if __name__ == "__main__":
    print("开始测试 Prajna Guangzhou Demo...")
    test_salary()
    test_sales_weekly()
    print("\n🎉 全部测试通过，Demo 可以正常运行。")
