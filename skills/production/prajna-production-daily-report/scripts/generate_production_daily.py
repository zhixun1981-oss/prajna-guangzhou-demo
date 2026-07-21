#!/usr/bin/env python3
"""
prajna-production-daily-report 生成器 v1.0.0
一键生成生产日报 Excel 套件：生产日报、产量统计、设备运行、质量检验、人员出勤。
"""
import argparse
import sys
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
SAMPLES_DIR = Path.home() / ".prajna" / "prajna-production-daily-report" / "samples"

BLUE = "1F4E78"
WHITE = "FFFFFF"


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )


def style_body(ws, start_row=2):
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 35)


def add_title(ws, title, merge_range):
    ws.merge_cells(merge_range)
    cell = ws[merge_range.split(":")[0]]
    cell.value = title
    cell.font = Font(bold=True, size=14, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def build_production_daily(args) -> Path:
    wb = Workbook()

    # Sheet 1: 生产日报
    ws1 = wb.active
    ws1.title = "生产日报"
    add_title(ws1, f"{args.factory} 生产日报（{args.date}）", "A1:F1")
    headers1 = ["车间", "生产线", "计划产量", "实际产量", "完成率", "差异原因", "当班负责人"]
    ws1.append([""] * len(headers1))
    ws1.append(headers1)
    style_header(ws1, row=2)
    sample1 = [
        ["组装车间", "A 线", 1200, 1150, "=D4/C4", "物料延迟 30 分钟", "李组长"],
        ["组装车间", "B 线", 1200, 1220, "=D5/C5", "效率提升", "王组长"],
        ["注塑车间", "1#机", 800, 820, "=D6/C6", "设备运行稳定", "张组长"],
        ["喷涂车间", "自动线", 600, 580, "=D7/C7", "环保检查停线 1 小时", "赵组长"],
    ]
    for row in sample1:
        ws1.append(row)
    style_body(ws1, start_row=3)
    auto_width(ws1)
    ws1.freeze_panes = "A3"

    # Sheet 2: 产量统计
    ws2 = wb.create_sheet("产量统计")
    add_title(ws2, f"{args.factory} 产量统计（近 7 日）", "A1:F1")
    headers2 = ["日期", "计划产量", "实际产量", "完成率", "累计产量", "备注"]
    ws2.append([""] * len(headers2))
    ws2.append(headers2)
    style_header(ws2, row=2)
    base = datetime.strptime(args.date, "%Y-%m-%d")
    cumulative = 0
    for i in range(6, -1, -1):
        d = (base - timedelta(days=i)).strftime("%Y-%m-%d")
        plan = 3600
        actual = plan + (50 if i % 2 == 0 else -80)
        cumulative += actual
        ws2.append([d, plan, actual, "=C{row}/B{row}".format(row=ws2.max_row + 1), cumulative, ""])
    style_body(ws2, start_row=3)
    auto_width(ws2)
    ws2.freeze_panes = "A3"

    # Sheet 3: 设备运行
    ws3 = wb.create_sheet("设备运行")
    add_title(ws3, f"{args.factory} 设备运行情况（{args.date}）", "A1:F1")
    headers3 = ["设备编号", "设备名称", "运行时长（h）", "故障时长（h）", "OEE", "故障描述", "维修状态"]
    ws3.append([""] * len(headers3))
    ws3.append(headers3)
    style_header(ws3, row=2)
    sample3 = [
        ["EQ-001", "注塑机 1#", 22, 2, "=C4/(C4+D4)", "液压泵异响", "已修复"],
        ["EQ-002", "注塑机 2#", 24, 0, "=C5/(C5+D5)", "无", "正常"],
        ["EQ-003", "自动喷涂线", 20, 4, "=C6/(C6+D6)", "环保检查停线", "待恢复"],
        ["EQ-004", "组装流水线 A", 23, 1, "=C7/(C7+D7)", "传送带卡料", "已修复"],
    ]
    for row in sample3:
        ws3.append(row)
    style_body(ws3, start_row=3)
    auto_width(ws3)
    ws3.freeze_panes = "A3"

    # Sheet 4: 质量检验
    ws4 = wb.create_sheet("质量检验")
    add_title(ws4, f"{args.factory} 质量检验日报（{args.date}）", "A1:F1")
    headers4 = ["工序", "检验数量", "合格数", "不合格数", "合格率", "主要不良项", "处理措施"]
    ws4.append([""] * len(headers4))
    ws4.append(headers4)
    style_header(ws4, row=2)
    sample4 = [
        ["来料检验", 500, 495, 5, "=C4/B4", "尺寸偏差", "退货供应商整改"],
        ["过程检验", 1200, 1185, 15, "=C5/B5", "外观划痕", "加强员工培训"],
        ["成品检验", 800, 792, 8, "=C6/B6", "功能测试不通过", "返工并追溯"],
    ]
    for row in sample4:
        ws4.append(row)
    style_body(ws4, start_row=3)
    auto_width(ws4)
    ws4.freeze_panes = "A3"

    # Sheet 5: 人员出勤
    ws5 = wb.create_sheet("人员出勤")
    add_title(ws5, f"{args.factory} 人员出勤（{args.date}）", "A1:F1")
    headers5 = ["车间", "应到人数", "实到人数", "出勤率", "缺勤人数", "缺勤原因", "加班人数"]
    ws5.append([""] * len(headers5))
    ws5.append(headers5)
    style_header(ws5, row=2)
    sample5 = [
        ["组装车间", 80, 78, "=C4/B4", 2, "病假 1 人，事假 1 人", 12],
        ["注塑车间", 45, 45, "=C5/B5", 0, "无", 8],
        ["喷涂车间", 30, 28, "=C6/B6", 2, "环保培训 2 人", 5],
        ["仓库", 20, 20, "=C7/B7", 0, "无", 3],
    ]
    for row in sample5:
        ws5.append(row)
    style_body(ws5, start_row=3)
    auto_width(ws5)
    ws5.freeze_panes = "A3"

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="生成生产日报 Excel 套件")
    parser.add_argument("--factory", default="智造工厂", help="工厂名称")
    parser.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"), help="日报日期")
    parser.add_argument("--output", "-o", help="输出文件路径")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    if not args.output:
        SAMPLES_DIR.mkdir(parents=True, exist_ok=True)
        args.output = SAMPLES_DIR / f"prajna_生产日报_{args.factory}_{args.date}.xlsx"
    path = build_production_daily(args)
    print(f"已生成生产日报套件：{path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
