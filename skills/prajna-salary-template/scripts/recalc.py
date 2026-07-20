#!/usr/bin/env python3
"""
prajna-salary-template 公式校验 / 轻量重算工具

用法：
    python3 recalc.py <file1.xlsx> [<file2.xlsx> ...]

功能：
1. 检查每个公式是否以 '=' 开头、括号是否平衡、所有工作表/单元格引用是否存在。
2. 按依赖顺序重算模板中的公式（支持 SUM / SUMPRODUCT / MAX / MIN / IF / AND /
   COUNTIF / COUNTA 以及基本算术比较运算符），并与 openpyxl data_only 缓存值比较。
3. 发现公式无效或数值不一致时返回非零退出码。

说明：
- 本工具只依赖 openpyxl；不需要 Excel 本身。
- 若文件从未被 Excel 打开保存过，data_only 缓存值可能为空，此时仅做语法与引用校验。
"""

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter


SUPPORTED_FUNCTIONS = {
    "SUM", "SUMPRODUCT", "MAX", "MIN", "IF", "AND",
    "COUNTIF", "COUNTA",
}

TWO_CHAR_OPS = {">=", "<=", "<>"}
OP_CHARS = set(">=<>&*/%+-(),!:")


def _looks_like_cell_or_range(val):
    """判断 val 是否为单元格/区域/整列/整行引用。"""
    if re.fullmatch(r"\$?[A-Za-z]+\$?\d*(:\$?[A-Za-z]+\$?\d*)?", val):
        return True
    if re.fullmatch(r"\$?\d+:\$?\d+", val):
        return True
    return False


def _normalize_ref(val):
    return val.replace("$", "")


def tokenize(formula):
    """将 Excel 公式拆分为 NUMBER / STRING / NAME / REF 标记。"""
    tokens = []
    i = 0
    n = len(formula)

    def is_ref_char(ch):
        return ch.isalnum() or ch == "_" or ch == "." or ("\u4e00" <= ch <= "\u9fff")

    while i < n:
        c = formula[i]
        if c.isspace():
            i += 1
            continue

        if c == '"':
            j = formula.find('"', i + 1)
            if j < 0:
                raise ValueError(f"未闭合的字符串：{formula[i:]}")
            tokens.append(("STRING", formula[i + 1:j]))
            i = j + 1
            continue

        if c == "'":
            j = formula.find("'", i + 1)
            if j < 0:
                raise ValueError(f"未闭合的工作表引号：{formula[i:]}")
            tokens.append(("STRING", formula[i + 1:j]))
            i = j + 1
            continue

        if i + 1 < n and formula[i:i + 2] in TWO_CHAR_OPS:
            tokens.append(("OP", formula[i:i + 2]))
            i += 2
            continue

        if c in OP_CHARS:
            tokens.append(("OP", c))
            i += 1
            continue

        if c == "$":
            i += 1
            c = formula[i] if i < n else ""

        if c.isdigit() or c == ".":
            j = i
            while j < n and (formula[j].isdigit() or formula[j] == "."):
                j += 1
            if j < n and formula[j] == "%":
                j += 1
            tokens.append(("NUMBER", formula[i:j]))
            i = j
            continue

        if is_ref_char(c):
            j = i
            while j < n and is_ref_char(formula[j]):
                j += 1
            val = formula[i:j]
            # 函数名或后面跟 ( 的标识符应为 NAME
            if j < n and formula[j] == "(":
                tokens.append(("NAME", val))
            elif _looks_like_cell_or_range(val):
                tokens.append(("REF", _normalize_ref(val)))
            else:
                tokens.append(("NAME", val))
            i = j
            continue

        raise ValueError(f"无法识别的字符 {c!r}（位置 {i}）")

    return _merge_ref_tokens(tokens)


def _merge_ref_tokens(tokens):
    """合并工作表前缀（NAME/STRING + ! + REF[/range]）和区域运算符（REF : REF）。"""
    def _is_reflike(t):
        return t[0] == "REF" or (t[0] == "NAME" and _looks_like_cell_or_range(t[1]))

    out = []
    i = 0
    n = len(tokens)
    while i < n:
        # sheet!ref[:ref]
        if (i + 2 < n
                and tokens[i][0] in ("NAME", "STRING")
                and tokens[i + 1] == ("OP", "!")
                and _is_reflike(tokens[i + 2])):
            sheet = tokens[i][1]
            ref = tokens[i + 2][1]
            i2 = i + 3
            if i2 + 1 < n and tokens[i2] == ("OP", ":") and _is_reflike(tokens[i2 + 1]):
                ref = f"{ref}:{tokens[i2 + 1][1]}"
                i2 += 2
            out.append(("REF", f"{sheet}!{ref}"))
            i = i2
            continue

        # ref:ref
        if (i + 2 < n
                and tokens[i][0] == "REF"
                and tokens[i + 1] == ("OP", ":")
                and _is_reflike(tokens[i + 2])):
            out.append(("REF", f"{tokens[i][1]}:{tokens[i + 2][1]}"))
            i += 3
            continue

        out.append(tokens[i])
        i += 1
    return out


def split_sheet(ref_token):
    if "!" in ref_token:
        sheet, rest = ref_token.split("!", 1)
        return sheet.strip().strip("'"), rest
    return None, ref_token


def _col_to_num(col):
    num = 0
    for ch in col.upper():
        num = num * 26 + (ord(ch) - ord("A") + 1)
    return num


def _iter_range(wb, sheet, rest):
    """遍历引用 token 涉及的所有 (sheet, coordinate) 坐标。"""
    if sheet is None:
        raise ValueError("引用缺少工作表")
    ws = wb[sheet]
    if ":" not in rest:
        yield sheet, rest.upper()
        return

    a, b = rest.split(":", 1)
    a = a.upper()
    b = b.upper()

    # 整列 F:F
    if a.isalpha() and b.isalpha():
        c1 = _col_to_num(a)
        c2 = _col_to_num(b)
        for c in range(c1, c2 + 1):
            for r in range(1, ws.max_row + 1):
                yield sheet, f"{get_column_letter(c)}{r}"
        return

    # 整行 3:3
    if a.isdigit() and b.isdigit():
        r1 = int(a)
        r2 = int(b)
        for r in range(r1, r2 + 1):
            for c in range(1, ws.max_column + 1):
                yield sheet, f"{get_column_letter(c)}{r}"
        return

    # 普通区域 A1:B10
    r1, c1 = coordinate_to_tuple(a)
    r2, c2 = coordinate_to_tuple(b)
    top, bottom = min(r1, r2), max(r1, r2)
    left, right = min(c1, c2), max(c1, c2)
    for r in range(top, bottom + 1):
        for c in range(left, right + 1):
            yield sheet, f"{get_column_letter(c)}{r}"


def validate_references(tokens, wb, current_sheet):
    """校验所有 REF token 引用的工作表和单元格是否真实存在。"""
    for typ, val in tokens:
        if typ != "REF":
            continue
        sheet, rest = split_sheet(val)
        sheet = sheet or current_sheet
        if sheet not in wb.sheetnames:
            raise ValueError(f"工作表不存在：{sheet}")
        ws = wb[sheet]
        if ":" not in rest:
            try:
                row, col = coordinate_to_tuple(rest)
            except Exception as exc:
                raise ValueError(f"无效单元格引用：{val}") from exc
            if row > ws.max_row or col > ws.max_column:
                raise ValueError(f"引用越界：{val}（工作表边界 {ws.max_row} 行 x {ws.max_column} 列）")
        else:
            a, b = rest.split(":", 1)
            # 整列/整行只做格式校验
            if a.isalpha() and b.isalpha():
                _col_to_num(a)
                _col_to_num(b)
            elif a.isdigit() and b.isdigit():
                int(a)
                int(b)
            else:
                try:
                    coordinate_to_tuple(a)
                    coordinate_to_tuple(b)
                except Exception as exc:
                    raise ValueError(f"无效区域引用：{val}") from exc


def _to_num(v):
    if v is None:
        return 0
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


def _flatten(args):
    for arg in args:
        if isinstance(arg, list):
            yield from _flatten(arg)
        else:
            yield arg


def _parse_criteria(criteria):
    s = str(criteria).strip()
    for op in (">=", "<=", "<>", ">", "<", "="):
        if s.startswith(op):
            val = s[len(op):].strip()
            try:
                val = float(val)
            except ValueError:
                pass
            return op, val
    try:
        return "=", float(s)
    except ValueError:
        return "=", s


def _compare(left, op, right):
    if op == ">=":
        return left >= right
    if op == "<=":
        return left <= right
    if op == ">":
        return left > right
    if op == "<":
        return left < right
    if op == "<>":
        return left != right
    return left == right


class Evaluator:
    def __init__(self, wb, stored_wb=None):
        self.wb = wb
        self.stored_wb = stored_wb
        self.cache = {}
        self.errors = []

    def _cell_value(self, sheet, coord):
        key = (sheet, coord)
        if key in self.cache:
            return self.cache[key]
        cell = self.wb[sheet][coord]
        if isinstance(cell.value, str) and cell.value.startswith("="):
            try:
                result = self._eval_formula(sheet, cell.value[1:])
            except Exception as exc:
                self.errors.append(f"{sheet}!{coord}: {exc}")
                result = None
        else:
            result = cell.value
        self.cache[key] = result
        return result

    def _resolve_ref(self, token, current_sheet):
        sheet, rest = split_sheet(token)
        sheet = sheet or current_sheet
        values = []
        for sh, coord in _iter_range(self.wb, sheet, rest):
            values.append(self._cell_value(sh, coord))
        if ":" not in rest and len(values) == 1:
            return values[0]
        return values

    def _call(self, name, args, current_sheet):
        name = name.upper()
        if name == "SUM":
            total = 0.0
            for v in _flatten(args):
                total += _to_num(v)
            return total
        if name == "SUMPRODUCT":
            arrays = []
            for arg in args:
                if not isinstance(arg, list):
                    arg = [arg]
                arrays.append([_to_num(x) for x in arg])
            if not arrays:
                return 0
            length = min(len(a) for a in arrays)
            if len(arrays) == 1:
                return sum(arrays[0][i] for i in range(length))
            total = 0.0
            for i in range(length):
                prod = 1.0
                for a in arrays:
                    prod *= a[i]
                total += prod
            return total
        if name in ("MAX", "MIN"):
            nums = [_to_num(v) for v in _flatten(args)]
            if not nums:
                return 0
            return max(nums) if name == "MAX" else min(nums)
        if name == "IF":
            if len(args) < 3:
                raise ValueError("IF 需要 3 个参数")
            cond = args[0]
            if isinstance(cond, list):
                cond = bool(cond[0]) if cond else False
            return args[1] if cond else args[2]
        if name == "AND":
            for v in _flatten(args):
                if not bool(v):
                    return False
            return True
        if name == "COUNTIF":
            if len(args) != 2:
                raise ValueError("COUNTIF 需要 2 个参数")
            values = args[0] if isinstance(args[0], list) else [args[0]]
            op, target = _parse_criteria(args[1])
            count = 0
            for v in values:
                if v is None:
                    continue
                if isinstance(target, str):
                    if op == "=" and str(v) == target:
                        count += 1
                else:
                    try:
                        if _compare(_to_num(v), op, target):
                            count += 1
                    except Exception:
                        pass
            return count
        if name == "COUNTA":
            values = list(_flatten(args))
            return sum(1 for v in values if v is not None)
        raise ValueError(f"不支持的函数：{name}")

    def _eval_formula(self, sheet, formula):
        tokens = tokenize(formula)
        result, pos = self._parse_expr(tokens, 0, sheet)
        if pos != len(tokens):
            raise ValueError(f"公式末尾有多余标记：{tokens[pos:]}")
        return result

    def _parse_expr(self, tokens, pos, sheet):
        return self._parse_concat(tokens, pos, sheet)

    def _parse_concat(self, tokens, pos, sheet):
        left, pos = self._parse_comparison(tokens, pos, sheet)
        while pos < len(tokens) and tokens[pos] == ("OP", "&"):
            pos += 1
            right, pos = self._parse_comparison(tokens, pos, sheet)
            left = str(left if left is not None else "") + str(right if right is not None else "")
        return left, pos

    def _parse_comparison(self, tokens, pos, sheet):
        left, pos = self._parse_additive(tokens, pos, sheet)
        if pos < len(tokens) and tokens[pos][0] == "OP" and tokens[pos][1] in (">=", "<=", "<>", ">", "<", "="):
            op = tokens[pos][1]
            pos += 1
            right, pos = self._parse_additive(tokens, pos, sheet)
            return _compare(_to_num(left), op, _to_num(right)), pos
        return left, pos

    def _parse_additive(self, tokens, pos, sheet):
        left, pos = self._parse_multiplicative(tokens, pos, sheet)
        while pos < len(tokens) and tokens[pos] in (("OP", "+"), ("OP", "-")):
            op = tokens[pos][1]
            pos += 1
            right, pos = self._parse_multiplicative(tokens, pos, sheet)
            left = (_to_num(left) + _to_num(right)) if op == "+" else (_to_num(left) - _to_num(right))
        return left, pos

    def _parse_multiplicative(self, tokens, pos, sheet):
        left, pos = self._parse_unary(tokens, pos, sheet)
        while pos < len(tokens) and tokens[pos] in (("OP", "*"), ("OP", "/")):
            op = tokens[pos][1]
            pos += 1
            right, pos = self._parse_unary(tokens, pos, sheet)
            if op == "*":
                left = _to_num(left) * _to_num(right)
            else:
                denom = _to_num(right)
                left = _to_num(left) / denom if denom != 0 else 0
        return left, pos

    def _parse_unary(self, tokens, pos, sheet):
        if pos < len(tokens) and tokens[pos] in (("OP", "+"), ("OP", "-")):
            op = tokens[pos][1]
            pos += 1
            val, pos = self._parse_unary(tokens, pos, sheet)
            return (+_to_num(val) if op == "+" else -_to_num(val)), pos
        val, pos = self._parse_primary(tokens, pos, sheet)
        while pos < len(tokens) and tokens[pos] == ("OP", "%"):
            pos += 1
            val = _to_num(val) / 100.0
        return val, pos

    def _parse_primary(self, tokens, pos, sheet):
        if pos >= len(tokens):
            raise ValueError("公式意外结束")
        typ, val = tokens[pos]
        pos += 1

        if typ == "NUMBER":
            if val.endswith("%"):
                return float(val[:-1]) / 100.0, pos
            return float(val), pos

        if typ == "STRING":
            return val, pos

        if typ == "REF":
            return self._resolve_ref(val, sheet), pos

        if typ == "NAME":
            if pos < len(tokens) and tokens[pos] == ("OP", "("):
                pos += 1
                args = []
                if pos < len(tokens) and tokens[pos] != ("OP", ")"):
                    arg, pos = self._parse_expr(tokens, pos, sheet)
                    args.append(arg)
                    while pos < len(tokens) and tokens[pos] == ("OP", ","):
                        pos += 1
                        arg, pos = self._parse_expr(tokens, pos, sheet)
                        args.append(arg)
                if pos >= len(tokens) or tokens[pos] != ("OP", ")"):
                    raise ValueError("函数调用缺少右括号")
                pos += 1
                return self._call(val, args, sheet), pos
            raise ValueError(f"未知标识符：{val}")

        if typ == "OP" and val == "(":
            result, pos = self._parse_expr(tokens, pos, sheet)
            if pos >= len(tokens) or tokens[pos] != ("OP", ")"):
                raise ValueError("缺少右括号")
            pos += 1
            return result, pos

        raise ValueError(f"无法解析的标记：{(typ, val)}")


def validate_formula(cell_value, sheet, wb):
    """校验公式语法和引用。返回 (ok, error_message)。"""
    if not isinstance(cell_value, str) or not cell_value.startswith("="):
        return False, "公式不以 '=' 开头"
    formula = cell_value[1:]
    depth = 0
    for ch in formula:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
            if depth < 0:
                return False, "括号不匹配（多余右括号）"
    if depth != 0:
        return False, "括号不匹配"
    try:
        tokens = tokenize(formula)
    except Exception as exc:
        return False, f"分词失败：{exc}"
    try:
        validate_references(tokens, wb, sheet)
    except Exception as exc:
        return False, f"引用校验失败：{exc}"
    return True, ""


def recalc_file(path):
    print(f"\n校验文件：{path}")
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as exc:
        print(f"  无法打开：{exc}")
        return 1

    try:
        stored_wb = load_workbook(path, data_only=True)
    except Exception:
        stored_wb = None

    evaluator = Evaluator(wb, stored_wb)
    formula_cells = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells.append((sheet, cell.coordinate))

    invalid = 0
    mismatches = 0
    errors = 0

    for sheet, coord in formula_cells:
        cell = wb[sheet][coord]
        ok, msg = validate_formula(cell.value, sheet, wb)
        if not ok:
            print(f"  INVALID {sheet}!{coord}: {msg}")
            invalid += 1
            continue

        try:
            computed = evaluator._cell_value(sheet, coord)
        except Exception as exc:
            print(f"  ERROR   {sheet}!{coord}: {exc}")
            errors += 1
            continue

        stored = None
        if stored_wb is not None:
            try:
                stored = stored_wb[sheet][coord].value
            except Exception:
                stored = None

        if stored is not None:
            if isinstance(computed, (int, float)) and isinstance(stored, (int, float)):
                if abs(computed - stored) > 0.01:
                    print(f"  MISMATCH {sheet}!{coord}: 计算值={computed}, 缓存值={stored}")
                    mismatches += 1
            elif str(computed) != str(stored):
                print(f"  MISMATCH {sheet}!{coord}: 计算值={computed!r}, 缓存值={stored!r}")
                mismatches += 1

    if evaluator.errors:
        for err in evaluator.errors:
            print(f"  EVAL-ERROR: {err}")
        errors += len(evaluator.errors)

    total = len(formula_cells)
    print(f"  公式数量：{total}，无效：{invalid}，不一致：{mismatches}，求值错误：{errors}")
    return 1 if (invalid or mismatches or errors) else 0


def main():
    parser = argparse.ArgumentParser(description="校验并重算 prajna 薪资模板 Excel 公式")
    parser.add_argument("paths", nargs="+", help="一个或多个 .xlsx 文件路径")
    args = parser.parse_args()

    rc = 0
    for p in args.paths:
        path = Path(p)
        if not path.exists():
            print(f"文件不存在：{path}")
            rc = 1
            continue
        rc = max(rc, recalc_file(str(path)))
    return rc


if __name__ == "__main__":
    sys.exit(main())
