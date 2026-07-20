#!/usr/bin/env python3
"""
prajna-compensation-system 生成器 v1.0.0
为企业一键生成完整的薪酬体系 Excel 工作簿。
包含岗位薪酬矩阵、薪酬结构、城市系数、调薪方案、预算、人力成本测算、绩效联动 7 大工作表。
与 prajna-salary-template 的职级带宽、城市系数、岗位类型打通，招聘/绩效技能可复用。
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
DATA_DIR = SKILL_DIR / "data"
DEFAULT_SAMPLES_DIR = Path.home() / ".prajna" / "prajna-compensation-system" / "samples"
SALARY_TEMPLATE_DATA_DIR = (
    Path.home() / ".prajna" / "skills" / "hr" / "prajna-salary-template" / "data"
)


# ---------------------------------------------------------------------------
# Embedded fallbacks (kept in sync with prajna-salary-template)
# ---------------------------------------------------------------------------
DEFAULT_LEVEL_BANDWIDTH = {
    "P1": {"name": "初级专员", "monthly_min": 3500, "monthly_mid": 4500, "monthly_max": 6000},
    "P2": {"name": "专员", "monthly_min": 5000, "monthly_mid": 6500, "monthly_max": 8500},
    "P3": {"name": "高级专员", "monthly_min": 7000, "monthly_mid": 9000, "monthly_max": 12000},
    "P4": {"name": "主管", "monthly_min": 9500, "monthly_mid": 12500, "monthly_max": 17000},
    "P5": {"name": "经理", "monthly_min": 14000, "monthly_mid": 19000, "monthly_max": 26000},
    "P6": {"name": "高级经理", "monthly_min": 22000, "monthly_mid": 30000, "monthly_max": 42000},
    "P7": {"name": "总监", "monthly_min": 35000, "monthly_mid": 48000, "monthly_max": 68000},
}

DEFAULT_POSITION_TYPES = {
    "技术类": {"base_ratio": 0.55, "position_ratio": 0.20, "performance_ratio": 0.15, "subsidy_ratio": 0.05, "bonus_ratio": 0.05},
    "产品类": {"base_ratio": 0.50, "position_ratio": 0.20, "performance_ratio": 0.20, "subsidy_ratio": 0.05, "bonus_ratio": 0.05},
    "运营类": {"base_ratio": 0.45, "position_ratio": 0.20, "performance_ratio": 0.20, "subsidy_ratio": 0.08, "bonus_ratio": 0.07},
    "销售类": {"base_ratio": 0.35, "position_ratio": 0.15, "performance_ratio": 0.20, "subsidy_ratio": 0.05, "bonus_ratio": 0.25},
    "生产类": {"base_ratio": 0.50, "position_ratio": 0.20, "performance_ratio": 0.15, "subsidy_ratio": 0.08, "bonus_ratio": 0.07},
    "质量类": {"base_ratio": 0.55, "position_ratio": 0.20, "performance_ratio": 0.12, "subsidy_ratio": 0.08, "bonus_ratio": 0.05},
    "供应链类": {"base_ratio": 0.50, "position_ratio": 0.20, "performance_ratio": 0.15, "subsidy_ratio": 0.08, "bonus_ratio": 0.07},
    "行政类": {"base_ratio": 0.55, "position_ratio": 0.20, "performance_ratio": 0.10, "subsidy_ratio": 0.08, "bonus_ratio": 0.07},
    "财务类": {"base_ratio": 0.50, "position_ratio": 0.20, "performance_ratio": 0.15, "subsidy_ratio": 0.08, "bonus_ratio": 0.07},
    "法务类": {"base_ratio": 0.50, "position_ratio": 0.20, "performance_ratio": 0.15, "subsidy_ratio": 0.08, "bonus_ratio": 0.07},
    "客服类": {"base_ratio": 0.50, "position_ratio": 0.20, "performance_ratio": 0.12, "subsidy_ratio": 0.08, "bonus_ratio": 0.10},
}

DEFAULT_CITY_MULTIPLIERS = {
    "一线城市": {"cities": ["北京", "上海", "深圳", "广州"], "multiplier": 1.30},
    "新一线城市": {"cities": ["杭州", "成都", "南京", "武汉", "重庆", "天津", "苏州", "西安", "长沙", "青岛", "郑州", "宁波", "东莞", "佛山", "合肥"], "multiplier": 1.15},
    "二线城市": {"cities": ["无锡", "厦门", "福州", "济南", "沈阳", "大连", "昆明", "哈尔滨", "长春", "石家庄", "南昌", "贵阳", "南宁", "海口", "兰州", "银川", "西宁", "乌鲁木齐", "呼和浩特", "太原"], "multiplier": 1.00},
    "三线城市": {"cities": ["珠海", "惠州", "中山", "南通", "常州", "徐州", "扬州", "绍兴", "嘉兴", "台州", "金华", "温州", "泉州", "烟台", "威海", "潍坊", "淄博", "临沂", "唐山", "秦皇岛", "保定", "邯郸", "洛阳", "南阳", "襄阳", "宜昌", "岳阳", "常德", "株洲", "湘潭", "柳州", "桂林", "三亚", "拉萨", "大庆", "包头", "鄂尔多斯", "芜湖", "蚌埠", "滁州", "马鞍山", "安庆", "阜阳", "宿州", "六安", "淮安", "盐城", "连云港", "宿迁", "镇江", "泰州", "湖州", "衢州", "舟山", "丽水"], "multiplier": 0.85},
    "四线及以下": {"cities": [], "multiplier": 0.75},
}


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------
def _load_json(path: Path, default):
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return default


def load_city_multipliers():
    data = _load_json(SALARY_TEMPLATE_DATA_DIR / "city_salary_multipliers.json", {})
    tier = data.get("tier", DEFAULT_CITY_MULTIPLIERS)
    return tier


def load_level_bandwidth():
    data = _load_json(SALARY_TEMPLATE_DATA_DIR / "job_level_bandwidth.json", {})
    levels = data.get("levels", {})
    if not levels:
        return DEFAULT_LEVEL_BANDWIDTH
    result = {}
    for level, info in levels.items():
        # Convert annual base salary from salary template into monthly matrix values.
        # Use base_salary + position_salary + performance_std as monthly total bandwidth proxy.
        base_min = info.get("base_salary_min", 0)
        base_mid = info.get("base_salary_mid", 0)
        base_max = info.get("base_salary_max", 0)
        pos_min = info.get("position_salary_min", 0)
        pos_mid = info.get("position_salary_mid", 0)
        pos_max = info.get("position_salary_max", 0)
        perf_min = info.get("performance_std_min", 0)
        perf_mid = info.get("performance_std_mid", 0)
        perf_max = info.get("performance_std_max", 0)
        result[level] = {
            "name": info.get("name", level),
            "monthly_min": base_min + pos_min + perf_min,
            "monthly_mid": base_mid + pos_mid + perf_mid,
            "monthly_max": base_max + pos_max + perf_max,
        }
    return result


def load_position_types():
    """
    加载岗位序列薪酬结构比例。
    若 prajna-salary-template 的 position_salary_templates.json 已包含 ratio 字段则直接复用；
    否则使用本技能内置的标准结构比例，确保各序列基本工资/岗位工资/绩效工资/津贴/奖金占比合理。
    """
    data = _load_json(SALARY_TEMPLATE_DATA_DIR / "position_salary_templates.json", {})
    types = data.get("position_types", {})
    normalized = {}
    for seq, vals in types.items():
        if "base_ratio" in vals:
            normalized[seq] = vals
    # Fallback: use embedded ratio presets for any missing sequence.
    for seq, ratios in DEFAULT_POSITION_TYPES.items():
        if seq not in normalized:
            normalized[seq] = ratios
    return normalized


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _safe_filename(name: str) -> str:
    name = re.sub(r"[\\/:*?\"<>|\s]+", "_", str(name))
    return name.strip("_") or "compensation"


def _today_cn() -> str:
    return datetime.now().strftime("%Y年%m月%d日")


def _today_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _get_city_multiplier(city: str, city_data: dict) -> float:
    for tier, info in city_data.items():
        if city in info.get("cities", []):
            return info.get("multiplier", 1.0)
    return 1.0


def _style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_subheader(cell):
    cell.font = Font(bold=True, color="000000")
    cell.fill = PatternFill("solid", fgColor="D9E1F2")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_note(cell):
    cell.font = Font(italic=True, color="666666", size=9)
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def _border_range(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border


def _write_title(ws, title, company, date):
    ws.append([f"{company} · {title}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.append([f"生成日期：{date}    版本：v1.0.0"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.append([])


def _autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
def build_sheet_matrix(ws, company, city, city_mult, level_bandwidth, position_types, seqs, levels):
    _write_title(ws, "岗位薪酬矩阵（岗位序列 × 职级）", company, _today_cn())
    headers = ["岗位序列", "职级", "职级名称", "城市", "城市系数", "月薪最小值", "月薪中位值", "月薪最大值", "年薪最小值", "年薪中位值", "年薪最大值", "备注"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        _style_header(ws.cell(row=ws.max_row, column=col))

    for seq in seqs:
        ratios = position_types.get(seq, DEFAULT_POSITION_TYPES.get(seq, DEFAULT_POSITION_TYPES["运营类"]))
        for level in levels:
            info = level_bandwidth.get(level, {})
            monthly_min = int(info.get("monthly_min", 0) * city_mult)
            monthly_mid = int(info.get("monthly_mid", 0) * city_mult)
            monthly_max = int(info.get("monthly_max", 0) * city_mult)
            note = f"结构占比：基本{int(ratios['base_ratio']*100)}%/岗位{int(ratios['position_ratio']*100)}%/绩效{int(ratios['performance_ratio']*100)}%/津贴{int(ratios['subsidy_ratio']*100)}%/奖金{int(ratios['bonus_ratio']*100)}%"
            ws.append([
                seq,
                level,
                info.get("name", ""),
                city,
                city_mult,
                monthly_min,
                monthly_mid,
                monthly_max,
                monthly_min * 12,
                monthly_mid * 12,
                monthly_max * 12,
                note,
            ])
    _border_range(ws, 4, ws.max_row, 1, len(headers))
    ws.append([])
    ws.append(["说明：本表月薪最小值/中位值/最大值基于 prajna-salary-template 职级带宽与城市系数测算，实际落地需结合岗位价值评估、市场薪酬调研及内部平衡结果进行人工复核。"])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))
    _style_note(ws.cell(row=ws.max_row, column=1))
    _autofit(ws)


def build_sheet_structure(ws, company, position_types, seqs):
    _write_title(ws, "薪酬结构表", company, _today_cn())
    headers = ["岗位序列", "基本工资占比", "岗位工资占比", "绩效工资占比", "津贴补贴占比", "奖金占比", "合计", "典型说明"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        _style_header(ws.cell(row=ws.max_row, column=col))

    descriptions = {
        "技术类": "基本工资高、绩效稳定，适合长期激励与项目奖金",
        "产品类": "岗位工资与绩效工资并重，强调交付质量与用户价值",
        "运营类": "绩效与奖金联动强，关注过程指标与结果指标双达成",
        "销售类": "低固定、高浮动，奖金（提成）占比高，结果导向",
        "生产类": "固定收入为主，绩效与质量、产量、安全挂钩",
        "质量类": "稳定性强，绩效与检验准确率、漏检率挂钩",
        "供应链类": "绩效与交付及时率、库存周转、采购降本挂钩",
        "行政类": "固定占比高，绩效侧重服务满意度与流程效率",
        "财务类": "固定占比高，绩效与报表质量、合规、风控挂钩",
        "法务类": "固定占比高，绩效与合同审核、风险处置挂钩",
        "客服类": "绩效与满意度、响应时效、解决率挂钩",
    }
    for seq in seqs:
        ratios = position_types.get(seq, DEFAULT_POSITION_TYPES.get(seq, DEFAULT_POSITION_TYPES["运营类"]))
        total = sum(ratios.values())
        ws.append([
            seq,
            f"{ratios['base_ratio']:.0%}",
            f"{ratios['position_ratio']:.0%}",
            f"{ratios['performance_ratio']:.0%}",
            f"{ratios['subsidy_ratio']:.0%}",
            f"{ratios['bonus_ratio']:.0%}",
            f"{total:.0%}",
            descriptions.get(seq, ""),
        ])
    _border_range(ws, 4, ws.max_row, 1, len(headers))
    ws.append([])
    ws.append(["说明：薪酬结构比例用于指导 offer 定薪、年度调薪与成本测算，不同序列可根据业务阶段与人才市场竞争度上下浮动 3-5 个百分点。"])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))
    _style_note(ws.cell(row=ws.max_row, column=1))
    _autofit(ws)


def build_sheet_city(ws, company, city_data):
    _write_title(ws, "城市薪酬系数", company, _today_cn())
    headers = ["城市层级", "城市", "薪酬系数", "社保基数下限", "社保基数上限", "公积金下限", "公积金上限", "备注"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        _style_header(ws.cell(row=ws.max_row, column=col))

    city_detail = _load_json(SALARY_TEMPLATE_DATA_DIR / "city_salary_multipliers.json", {}).get("city_detail", {})
    for tier, info in city_data.items():
        for city in info.get("cities", []):
            detail = city_detail.get(city, {})
            ws.append([
                tier,
                city,
                info.get("multiplier", 1.0),
                detail.get("social_base_min", "-"),
                detail.get("social_base_max", "-"),
                detail.get("housing_base_min", "-"),
                detail.get("housing_base_max", "-"),
                "",
            ])
    _border_range(ws, 4, ws.max_row, 1, len(headers))
    ws.append([])
    ws.append(["说明：城市系数用于统一不同城市同岗同级的薪酬基准，招聘 offer 定薪、年度预算、异地调岗均以此表为参考。数据与 prajna-salary-template 保持一致。"])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))
    _style_note(ws.cell(row=ws.max_row, column=1))
    _autofit(ws)


def build_sheet_raise_plan(ws, company, level_bandwidth, seqs, levels, growth):
    _write_title(ws, "调薪方案", company, _today_cn())
    headers = ["调薪类型", "适用对象", "调薪幅度", "生效周期", "预算占比", "审批层级", "操作要点"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        _style_header(ws.cell(row=ws.max_row, column=col))

    rows = [
        ["普调", "全体员工", f"{growth}%", "次年 1 月", "40%", "CEO/HRD", "参考 CPI、市场薪酬涨幅、公司营收增速，统一调整固定薪酬基数。"],
        ["晋升调薪", "通过晋升评估的员工", "15%-25%", "晋升生效月", "30%", "部门负责人+HRD", "依据岗位薪酬矩阵，按新职级中位值 80%-100% 定薪，避免薪酬倒挂。"],
        ["绩效调薪", "绩效结果前 20% 员工", "5%-15%", "次年 1 月", "20%", "部门负责人+HR", "与年度绩效等级挂钩，A 级 10%-15%，B+ 级 5%-10%，C/D 级不调。"],
        ["特殊保留调薪", "关键岗位/高潜/市场稀缺人才", "10%-30%", "随时", "10%", "CEO/HRD", "需提交人才保留申请，说明市场缺口、替代成本、业务影响与 ROI。"],
        ["试用期转正", "试用期员工", "0%-8%", "转正月", "不占年度预算", "部门负责人+HR", "依据试用期考核结果，可在薪酬带宽内小幅上调。"],
    ]
    for row in rows:
        ws.append(row)
    _border_range(ws, 4, ws.max_row, 1, len(headers))
    ws.append([])
    ws.append(["说明：调薪总额应控制在年度薪酬预算的合理范围内，特殊保留调薪需建立专项台账，避免预算超支与内部公平性失衡。"])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))
    _style_note(ws.cell(row=ws.max_row, column=1))
    _autofit(ws)


def build_sheet_budget(ws, company, industry, scale, budget, city, city_mult, level_bandwidth, seqs, levels):
    _write_title(ws, "薪酬预算表（年度人力成本预算）", company, _today_cn())
    headers = ["部门/序列", "职级", "人数", "月薪中位值", "年度固定成本", "年度奖金/绩效", "社保公积金（企业部分）", "其他福利", "年度总成本", "占总预算比例"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        _style_header(ws.cell(row=ws.max_row, column=col))

    total_headcount = scale
    # Distribute headcount evenly across seq × level combinations.
    combinations = [(seq, level) for seq in seqs for level in levels]
    per_combo = max(1, total_headcount // len(combinations))
    remainder = total_headcount - per_combo * len(combinations)

    grand_total = 0
    rows = []
    for idx, (seq, level) in enumerate(combinations):
        info = level_bandwidth.get(level, {})
        monthly_mid = int(info.get("monthly_mid", 0) * city_mult)
        headcount = per_combo + (1 if idx < remainder else 0)
        annual_fixed = monthly_mid * 12 * headcount
        annual_bonus = int(annual_fixed * 0.15)  # 15% bonus proxy
        social_cost = int((monthly_mid * 0.30) * 12 * headcount)  # employer social ~30%
        other_welfare = int((monthly_mid * 0.05) * 12 * headcount)
        total = annual_fixed + annual_bonus + social_cost + other_welfare
        grand_total += total
        rows.append([seq, level, headcount, monthly_mid, annual_fixed, annual_bonus, social_cost, other_welfare, total, 0.0])

    for row in rows:
        row[9] = round(row[8] / grand_total, 4) if grand_total else 0
        ws.append(row)

    _border_range(ws, 4, ws.max_row, 1, len(headers))
    last_data_row = ws.max_row
    ws.append([])
    ws.append(["合计", "", total_headcount, "", "=SUM(E5:E" + str(last_data_row) + ")", "=SUM(F5:F" + str(last_data_row) + ")", "=SUM(G5:G" + str(last_data_row) + ")", "=SUM(H5:H" + str(last_data_row) + ")", "=SUM(I5:I" + str(last_data_row) + ")", "=SUM(J5:J" + str(last_data_row) + ")"])
    _border_range(ws, ws.max_row, ws.max_row, 1, len(headers))
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append([])
    ws.append([f"说明：本表按 {industry} 行业、{scale} 人规模、总部 {city}（系数 {city_mult}）测算，年度总预算目标 {budget:,.0f} 元。实际编制应结合组织架构、HC 计划与业务节奏细化。"])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))
    _style_note(ws.cell(row=ws.max_row, column=1))
    _autofit(ws)


def build_sheet_cost(ws, company, city, level_bandwidth, seqs, levels, city_mult):
    _write_title(ws, "人力成本测算", company, _today_cn())
    headers = ["成本项目", "计算基数", "企业承担比例/说明", "月度人均成本", "年度人均成本", "测算示例（月薪中位值）", "备注"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        _style_header(ws.cell(row=ws.max_row, column=col))

    sample_mid = int(level_bandwidth.get("P3", {}).get("monthly_mid", 9000) * city_mult)
    rows = [
        ["养老保险", "社保基数", "16%", f"=G5*0.16", f"=E5*12", sample_mid, "基数上下限见城市系数表"],
        ["医疗保险", "社保基数", "8%-10%", f"=G6*0.09", f"=E6*12", sample_mid, "各地比例略有差异"],
        ["失业保险", "社保基数", "0.5%", f"=G7*0.005", f"=E7*12", sample_mid, ""],
        ["工伤保险", "社保基数", "0.2%-1.9%", f"=G8*0.005", f"=E8*12", sample_mid, "按行业风险等级"],
        ["生育保险", "社保基数", "0.5%-1%", f"=G9*0.008", f"=E9*12", sample_mid, "部分城市并入医保"],
        ["住房公积金", "公积金基数", "5%-12%", f"=G10*0.12", f"=E10*12", sample_mid, "企业与个人同比例"],
        ["社保公积金小计", "", "约 30%-42%", "=SUM(E5:E10)", "=SUM(F5:F10)", sample_mid, "企业承担部分"],
        ["年终奖", "月薪 × 月数", "1-3 个月", f"=G11*1.5", f"=E11*1.5", sample_mid, "按绩效结果浮动"],
        ["离职成本", "离职补偿 + 招聘 + 培训", "约 0.5-1.5 倍年薪", f"=G12*0.08", f"=E12*12*0.08", sample_mid, "离职率 8% 假设"],
        ["个税（代扣代缴）", "应税工资", "累进税率 3%-45%", "-", "-", sample_mid, "个人所得税由员工承担，企业代扣代缴"],
    ]
    for row in rows:
        ws.append(row)
    _border_range(ws, 4, ws.max_row, 1, len(headers))
    ws.append([])
    ws.append(["说明：社保公积金比例以当地最新政策为准，本表仅作测算参考；实际人力成本还应包含招聘费、培训费、办公费、商业保险、体检、团建等。"])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))
    _style_note(ws.cell(row=ws.max_row, column=1))
    _autofit(ws)


def build_sheet_performance_link(ws, company, seqs):
    _write_title(ws, "薪酬与绩效联动规则", company, _today_cn())
    headers = ["绩效等级", "绩效系数", "年终奖系数", "调薪幅度", "晋升资格", "适用说明"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        _style_header(ws.cell(row=ws.max_row, column=col))

    rows = [
        ["S（卓越）", "1.20-1.50", "2.5-4 个月", "15%-25%", "优先", "业绩显著超出预期，对公司有突出贡献"],
        ["A（优秀）", "1.05-1.20", "1.5-2.5 个月", "8%-15%", "具备", "持续高绩效，关键目标全面达成"],
        ["B+（良好）", "0.95-1.05", "1-1.5 个月", "3%-8%", "视岗位", "目标基本达成，部分超出预期"],
        ["B（合格）", "0.80-0.95", "0.5-1 个月", "0%-3%", "暂无", "基本达成岗位要求，需持续提升"],
        ["C（待改进）", "0.50-0.80", "0", "0", "无", "未达成关键目标，进入绩效改进计划 PIP"],
        ["D（不合格）", "0-0.50", "0", "0", "无", "严重未达标，按公司制度处理"],
    ]
    for row in rows:
        ws.append(row)
    _border_range(ws, 4, ws.max_row, 1, len(headers))

    ws.append([])
    ws.append(["岗位序列绩效权重参考"])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)

    seq_headers = ["岗位序列", "业绩结果权重", "过程/能力权重", "价值观/协作权重", "考核周期", "典型 KPI 方向"]
    ws.append(seq_headers)
    for col, _ in enumerate(seq_headers, 1):
        _style_subheader(ws.cell(row=ws.max_row, column=col))

    seq_rows = [
        ["销售类", "60%", "25%", "15%", "月度/季度", "销售额、回款率、客户开发、客单价"],
        ["运营类", "50%", "35%", "15%", "月度/季度", "GMV、转化率、活动 ROI、库存周转"],
        ["技术类", "40%", "45%", "15%", "季度", "交付质量、项目进度、系统稳定性、技术创新"],
        ["产品类", "45%", "40%", "15%", "季度", "需求交付、用户满意度、项目上线质量"],
        ["生产类", "50%", "35%", "15%", "月度", "产量、合格率、安全事故、成本控制"],
        ["质量类", "45%", "40%", "15%", "月度", "检验准确率、漏检率、异常反馈及时率"],
        ["供应链类", "50%", "35%", "15%", "季度", "交付及时率、库存周转、采购降本"],
        ["财务类", "40%", "45%", "15%", "季度/年度", "报表质量、合规、预算执行、风控"],
        ["行政/法务/客服", "35%", "45%", "20%", "季度", "服务满意度、流程效率、风险处置、响应时效"],
    ]
    for row in seq_rows:
        ws.append(row)
    _border_range(ws, ws.max_row - len(seq_rows), ws.max_row, 1, len(seq_headers))

    ws.append([])
    ws.append(["说明：绩效结果决定绩效奖金、年终奖与调薪资格，绩效系数用于计算绩效工资发放比例。具体权重与指标可由 prajna-performance-system 的 KPI 指标库进一步细化。"])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))
    _style_note(ws.cell(row=ws.max_row, column=1))
    _autofit(ws)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="生成企业薪酬体系 Excel 工作簿"
    )
    parser.add_argument(
        "--company", "-c",
        help="企业名称",
        default="智云科技",
    )
    parser.add_argument(
        "--industry",
        help="行业",
        default="互联网",
    )
    parser.add_argument(
        "--city",
        help="总部城市",
        default="广州",
    )
    parser.add_argument(
        "--scale",
        type=int,
        help="企业规模（人数）",
        default=200,
    )
    parser.add_argument(
        "--seqs",
        help="岗位序列，逗号分隔，默认：技术类,产品类,运营类,销售类,生产类,质量类,供应链类,行政类,财务类,法务类,客服类",
        default="技术类,产品类,运营类,销售类,生产类,质量类,供应链类,行政类,财务类,法务类,客服类",
    )
    parser.add_argument(
        "--levels",
        help="职级序列，逗号分隔，默认：P1,P2,P3,P4,P5,P6,P7",
        default="P1,P2,P3,P4,P5,P6,P7",
    )
    parser.add_argument(
        "--budget",
        type=float,
        help="年度薪酬总预算（元）",
        default=30000000,
    )
    parser.add_argument(
        "--growth",
        type=float,
        help="年度普调幅度（%%）",
        default=5.0,
    )
    parser.add_argument(
        "--output", "-o",
        help="输出文件路径，覆盖默认路径",
    )
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    seqs = [s.strip() for s in args.seqs.split(",") if s.strip()]
    levels = [l.strip() for l in args.levels.split(",") if l.strip()]

    city_data = load_city_multipliers()
    level_bandwidth = load_level_bandwidth()
    position_types = load_position_types()
    city_mult = _get_city_multiplier(args.city, city_data)

    # Filter to requested levels only, preserving order.
    level_bandwidth = {k: v for k, v in level_bandwidth.items() if k in levels}

    wb = Workbook()
    # Remove default sheet and rebuild in order.
    wb.remove(wb.active)

    build_sheet_matrix(wb.create_sheet("岗位薪酬矩阵"), args.company, args.city, city_mult, level_bandwidth, position_types, seqs, levels)
    build_sheet_structure(wb.create_sheet("薪酬结构表"), args.company, position_types, seqs)
    build_sheet_city(wb.create_sheet("城市薪酬系数"), args.company, city_data)
    build_sheet_raise_plan(wb.create_sheet("调薪方案"), args.company, level_bandwidth, seqs, levels, args.growth)
    build_sheet_budget(wb.create_sheet("薪酬预算表"), args.company, args.industry, args.scale, args.budget, args.city, city_mult, level_bandwidth, seqs, levels)
    build_sheet_cost(wb.create_sheet("人力成本测算"), args.company, args.city, level_bandwidth, seqs, levels, city_mult)
    build_sheet_performance_link(wb.create_sheet("薪酬与绩效联动规则"), args.company, seqs)

    # Disclaimer sheet
    ws = wb.create_sheet("使用说明与免责声明")
    ws.append(["使用说明与免责声明"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.append([])
    notes = [
        "1. 本薪酬体系工作簿由 Prajna 企智自动生成，用于快速搭建薪酬框架与预算测算。",
        "2. 岗位薪酬矩阵数据与 prajna-salary-template 的职级带宽、城市系数保持一致。",
        "3. 薪酬结构表中的比例可用于 prajna-recruitment-assistant 的 Offer 薪资建议。",
        "4. 绩效联动规则可与 prajna-performance-system 的 KPI 指标库、绩效评分表配套使用。",
        "5. 社保公积金比例、个税政策、最低工资标准请以当地最新法规为准，并由 HR 与财务人工复核。",
        "6. 本表不构成法律、税务或劳动用工建议，企业应根据自身情况调整并经法务/HR 审核后执行。",
    ]
    for note in notes:
        ws.append([note])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=3)
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    _autofit(ws)

    if args.output:
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
    else:
        DEFAULT_SAMPLES_DIR.mkdir(parents=True, exist_ok=True)
        safe_company = _safe_filename(args.company)
        filename = f"prajna_薪酬体系_{safe_company}_{args.industry}_{args.city}_{_today_iso()}.xlsx"
        output_path = DEFAULT_SAMPLES_DIR / filename

    wb.save(output_path)
    print(f"已生成薪酬体系工作簿：{output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
