#!/usr/bin/env python3
"""
prajna-customer-service-sop 生成器 v1.0.0
一键生成客服标准作业程序 Excel 套件：话术库、工单处理流程、客诉分类与升级规则、FAQ。
"""
import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
SAMPLES_DIR = Path.home() / ".prajna" / "prajna-customer-service-sop" / "samples"

BLUE = "1F4E78"
WHITE = "FFFFFF"


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )


def style_body(ws, start_row=2):
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 45)


def add_title(ws, title, merge_range):
    ws.merge_cells(merge_range)
    cell = ws[merge_range.split(":")[0]]
    cell.value = title
    cell.font = Font(bold=True, size=14, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def build_cs_sop(args) -> Path:
    wb = Workbook()

    # Sheet 1: 客服话术库
    ws1 = wb.active
    ws1.title = "客服话术库"
    add_title(ws1, f"{args.company} 客服话术库", "A1:E1")
    headers1 = ["场景", "客户情绪", "客服话术", "禁止用语", "备注"]
    ws1.append([""] * len(headers1))
    ws1.append(headers1)
    style_header(ws1, row=2)
    sample1 = [
        ["开场白", "中性", "您好，欢迎联系{company}客服，我是客服小 P，很高兴为您服务，请问有什么可以帮您？".format(company=args.company), "喂/什么事/说", "保持热情、语速适中"],
        ["物流查询", "焦急", "我理解您迫切希望收到商品的心情，请提供订单号，我立即为您查询物流最新状态。", "我不知道/你自己查", "主动安抚，先查后答"],
        ["退换货", "不满", "非常抱歉给您带来不好的体验，我们支持 7 天无理由退换货，我先帮您核实订单信息。", "不能退/你自己负责", "先道歉，再处理"],
        ["投诉产品质量", "愤怒", "非常抱歉产品未能达到您的期望，我们会认真对待您的反馈，请描述具体问题，我为您记录并升级。", "产品没问题/是你不会用", "记录详细，承诺时效"],
        ["结束语", "中性", "感谢您的来电/咨询，祝您生活愉快，如有其他问题随时联系我们。", "挂了/不说了", "确认客户无其他问题"],
    ]
    for row in sample1:
        ws1.append(row)
    style_body(ws1, start_row=3)
    auto_width(ws1)
    ws1.freeze_panes = "A3"

    # Sheet 2: 工单处理流程
    ws2 = wb.create_sheet("工单处理流程")
    add_title(ws2, f"{args.company} 工单处理流程", "A1:E1")
    headers2 = ["步骤", "处理环节", "处理人", "标准时效", "关键动作", "输出结果"]
    ws2.append([""] * len(headers2))
    ws2.append(headers2)
    style_header(ws2, row=2)
    sample2 = [
        [1, "工单接入", "客服代表", "5 分钟", "记录客户问题、情绪、联系方式", "生成标准化工单"],
        [2, "问题分类", "客服代表", "10 分钟", "按客诉分类标准打标签", "确定处理路径"],
        [3, "一线处理", "客服代表", "2 小时", "依据话术库与 FAQ 解答或处理", "客户确认解决"],
        [4, "升级二线", "客服主管", "4 小时", "复杂/高风险问题移交专业团队", "明确责任人与 SLA"],
        [5, "协同处理", "技术/物流/售后", "24 小时", "跨部门协作解决根因", "反馈处理结果"],
        [6, "客户回访", "客服代表", "48 小时", "确认客户满意度并关闭工单", "满意度评分≥4 分"],
    ]
    for row in sample2:
        ws2.append(row)
    style_body(ws2, start_row=3)
    auto_width(ws2)
    ws2.freeze_panes = "A3"

    # Sheet 3: 客诉分类与升级规则
    ws3 = wb.create_sheet("客诉分类与升级规则")
    add_title(ws3, f"{args.company} 客诉分类与升级规则", "A1:F1")
    headers3 = ["客诉类型", "典型表现", "一线处理权限", "升级条件", "升级对象", "处理时限"]
    ws3.append([""] * len(headers3))
    ws3.append(headers3)
    style_header(ws3, row=2)
    sample3 = [
        ["咨询类", "产品使用、政策询问", "全额解答", "无法解答或客户要求", "客服主管", "2 小时"],
        ["售后类", "退换货、维修、补偿", "按标准执行", "金额>500 元或客户情绪激动", "售后经理", "4 小时"],
        ["物流类", "未收到、破损、延误", "协调物流", "贵重物品丢失/破损", "物流经理", "24 小时"],
        ["质量类", "产品缺陷、安全隐患", "记录并安抚", "涉及人身安全或批量问题", "质量总监", "2 小时"],
        ["舆情类", "威胁投诉媒体/平台", "立即升级", "出现差评、社交媒体曝光", "公关/客服总监", "1 小时"],
    ]
    for row in sample3:
        ws3.append(row)
    style_body(ws3, start_row=3)
    auto_width(ws3)
    ws3.freeze_panes = "A3"

    # Sheet 4: FAQ
    ws4 = wb.create_sheet("FAQ")
    add_title(ws4, f"{args.company} 常见问题 FAQ", "A1:D1")
    headers4 = ["问题分类", "问题", "标准答案", "关联话术"]
    ws4.append([""] * len(headers4))
    ws4.append(headers4)
    style_header(ws4, row=2)
    sample4 = [
        ["账户", "如何修改收货地址？", "登录 App → 我的 → 地址管理 → 编辑/新增地址，下单前请确认默认地址。", "开场白"],
        ["订单", "订单什么时候发货？", "付款后 24 小时内发货，预售商品以页面标注时间为准。", "物流查询"],
        ["售后", "退换货运费谁承担？", "质量问题我们承担运费，非质量问题由买家承担，具体以售后判定为准。", "退换货"],
        ["支付", "支持哪些支付方式？", "支持微信支付、支付宝、银联卡及企业公对公转账。", "开场白"],
        ["发票", "如何申请开票？", "订单完成后 30 天内可在订单详情页申请电子发票，企业用户请联系客服。", "开场白"],
    ]
    for row in sample4:
        ws4.append(row)
    style_body(ws4, start_row=3)
    auto_width(ws4)
    ws4.freeze_panes = "A3"

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="生成客服 SOP Excel 套件")
    parser.add_argument("--company", default="智服电商", help="公司名称")
    parser.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"), help="更新日期")
    parser.add_argument("--output", "-o", help="输出文件路径")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    if not args.output:
        SAMPLES_DIR.mkdir(parents=True, exist_ok=True)
        args.output = SAMPLES_DIR / f"prajna_客服SOP_{args.company}_{args.date}.xlsx"
    path = build_cs_sop(args)
    print(f"已生成客服 SOP 套件：{path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
