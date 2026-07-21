#!/usr/bin/env python3
"""
prajna-procurement-assistant 生成器 v1.0.0
一键生成采购管理 Excel 套件：采购申请单、供应商评估表、询价比价单、
采购合同审查清单、采购台账。
"""
import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
SAMPLES_DIR = Path.home() / ".prajna" / "prajna-procurement-assistant" / "samples"

BLUE = "1F4E78"
LIGHT_BLUE = "D9E1F2"
WHITE = "FFFFFF"
GRAY = "F2F2F2"


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )


def style_body(ws, start_row=2):
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)


def add_title(ws, title, merge_range):
    ws.merge_cells(merge_range)
    cell = ws[merge_range.split(":")[0]]
    cell.value = title
    cell.font = Font(bold=True, size=14, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def build_procurement_kit(args) -> Path:
    wb = Workbook()

    # Sheet 1: 采购申请单
    ws1 = wb.active
    ws1.title = "采购申请单"
    add_title(ws1, f"{args.company} 采购申请单", "A1:F1")
    headers1 = ["申请日期", "申请部门", "申请人", "物料名称", "规格型号", "数量", "预算单价（元）", "用途说明", "期望到货日期", "审批状态"]
    ws1.append([""] * len(headers1))
    ws1.append(headers1)
    style_header(ws1, row=2)
    sample1 = [
        [args.date, args.department, args.applicant, "办公电脑", "14寸/i5/16G/512G", 5, 4800, "新员工入职配置", args.delivery_date, "待审批"],
        [args.date, args.department, args.applicant, "A4 打印纸", "70g/500张/包", 50, 25, "日常办公耗材", args.delivery_date, "待审批"],
        [args.date, args.department, args.applicant, "服务器硬盘", "2TB SAS 企业级", 4, 1800, "存储扩容", args.delivery_date, "待审批"],
    ]
    for row in sample1:
        ws1.append(row)
    style_body(ws1, start_row=3)
    auto_width(ws1)
    ws1.freeze_panes = "A3"

    # Sheet 2: 供应商评估表
    ws2 = wb.create_sheet("供应商评估表")
    add_title(ws2, f"{args.company} 供应商评估表", "A1:G1")
    headers2 = ["供应商名称", "供应品类", "注册资本（万元）", "合作年限", "质量评分", "交期评分", "价格评分", "服务评分", "综合得分", "评级", "是否合格"]
    ws2.append([""] * len(headers2))
    ws2.append(headers2)
    style_header(ws2, row=2)
    sample2 = [
        ["广州智信科技", "IT 设备", 500, 3, 92, 88, 85, 90, "=AVERAGE(E4:H4)", "=IF(I4>=90,'A',IF(I4>=80,'B','C'))", "=IF(I4>=80,'是','否')"],
        ["深圳优采供应链", "办公耗材", 300, 2, 85, 90, 95, 88, "=AVERAGE(E5:H5)", "=IF(I5>=90,'A',IF(I5>=80,'B','C'))", "=IF(I5>=80,'是','否')"],
        ["上海华云数据", "服务器配件", 2000, 5, 95, 85, 80, 92, "=AVERAGE(E6:H6)", "=IF(I6>=90,'A',IF(I6>=80,'B','C'))", "=IF(I6>=80,'是','否')"],
    ]
    for row in sample2:
        ws2.append(row)
    style_body(ws2, start_row=3)
    auto_width(ws2)
    ws2.freeze_panes = "A3"

    # Sheet 3: 询价比价单
    ws3 = wb.create_sheet("询价比价单")
    add_title(ws3, f"{args.company} 询价比价单", "A1:H1")
    headers3 = ["物料名称", "规格型号", "数量", "供应商A", "供应商A报价", "供应商B", "供应商B报价", "供应商C", "供应商C报价", "推荐供应商", "节省金额（元）"]
    ws3.append([""] * len(headers3))
    ws3.append(headers3)
    style_header(ws3, row=2)
    sample3 = [
        ["办公电脑", "14寸/i5/16G/512G", 5, "广州智信科技", 4800, "深圳优采供应链", 4950, "北京数码港", 5100, "=IF(E4<G4,IF(E4<I4,D4,F4),IF(G4<I4,F4,H4))", "=MAX(E4,G4,I4)-MIN(E4,G4,I4)"],
        ["A4 打印纸", "70g/500张/包", 50, "广州智信科技", 25, "深圳优采供应链", 23, "北京数码港", 24, "=IF(E5<G5,IF(E5<I5,D5,F5),IF(G5<I5,F5,H5))", "=MAX(E5,G5,I5)-MIN(E5,G5,I5)"],
    ]
    for row in sample3:
        ws3.append(row)
    style_body(ws3, start_row=3)
    auto_width(ws3)
    ws3.freeze_panes = "A3"

    # Sheet 4: 采购合同审查清单
    ws4 = wb.create_sheet("采购合同审查清单")
    add_title(ws4, f"{args.company} 采购合同审查清单", "A1:E1")
    headers4 = ["审查项目", "审查要点", "是否合规", "风险等级", "备注"]
    ws4.append([""] * len(headers4))
    ws4.append(headers4)
    style_header(ws4, row=2)
    sample4 = [
        ["合同主体", "供应商营业执照、资质、签约代表授权", "是", "高", "已核验三证合一"],
        ["标的条款", "物料名称、规格、数量、单价、总价清晰", "是", "高", "与采购申请一致"],
        ["交付条款", "交货时间、地点、运输方式、验收标准", "是", "中", "需明确到货破损责任"],
        ["付款条款", "付款比例、账期、发票类型、违约金", "是", "高", "30%预付，70%到货验收后 30 天"],
        ["质保售后", "质保期、退换货、售后服务响应时间", "是", "中", "整机 3 年质保"],
        ["保密与知识产权", "保密义务、知识产权归属", "是", "低", "标准条款"],
        ["争议解决", "管辖法院/仲裁机构、法律适用", "是", "中", "约定甲方所在地法院"],
    ]
    for row in sample4:
        ws4.append(row)
    style_body(ws4, start_row=3)
    auto_width(ws4)
    ws4.freeze_panes = "A3"

    # Sheet 5: 采购台账
    ws5 = wb.create_sheet("采购台账")
    add_title(ws5, f"{args.company} 采购台账", "A1:H1")
    headers5 = ["采购编号", "采购日期", "物料名称", "供应商", "数量", "单价（元）", "总价（元）", "付款状态", "到货状态", "发票状态"]
    ws5.append([""] * len(headers5))
    ws5.append(headers5)
    style_header(ws5, row=2)
    sample5 = [
        [f"CG-{args.date[:4]}{args.date[5:7]}001", args.date, "办公电脑", "广州智信科技", 5, 4800, "=E4*F4", "已付款", "已到货", "已收票"],
        [f"CG-{args.date[:4]}{args.date[5:7]}002", args.date, "A4 打印纸", "深圳优采供应链", 50, 23, "=E5*F5", "已付款", "已到货", "已收票"],
        [f"CG-{args.date[:4]}{args.date[5:7]}003", args.date, "服务器硬盘", "上海华云数据", 4, 1800, "=E6*F6", "待付款", "运输中", "未收票"],
    ]
    for row in sample5:
        ws5.append(row)
    style_body(ws5, start_row=3)
    auto_width(ws5)
    ws5.freeze_panes = "A3"

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="生成采购管理 Excel 套件")
    parser.add_argument("--company", default="智采科技", help="公司名称")
    parser.add_argument("--department", default="采购部", help="申请部门")
    parser.add_argument("--applicant", default="张采购", help="申请人")
    parser.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"), help="申请日期")
    parser.add_argument("--delivery-date", default=datetime.now().strftime("%Y-%m-%d"), help="期望到货日期")
    parser.add_argument("--output", "-o", help="输出文件路径")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    if not args.output:
        SAMPLES_DIR.mkdir(parents=True, exist_ok=True)
        args.output = SAMPLES_DIR / f"prajna_采购管理套件_{args.company}_{args.date}.xlsx"
    path = build_procurement_kit(args)
    print(f"已生成采购管理套件：{path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
