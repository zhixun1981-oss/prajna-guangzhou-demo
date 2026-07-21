#!/usr/bin/env python3
"""
prajna-sales-weekly-report 生成器 v1.0.0
为销售团队一键生成标准周报 Excel 模板。
支持：团队预设、自定义目标、周期、输出路径；内置 5 张工作表与核心 KPI 公式。
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
DATA_DIR = SKILL_DIR / "data"
HISTORY_DIR = Path.home() / ".prajna" / "sales_weekly_report_history"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _safe_filename(name):
    name = re.sub(r"[\\/:*?\"<>|\s]+", "_", str(name))
    return name.strip("_") or "team"


def _base_styles():
    return {
        "title_font": Font(name="Arial", bold=True, size=16, color="1F4E78"),
        "section_font": Font(name="Arial", bold=True, size=12, color="1F4E78"),
        "header_font": Font(name="Arial", bold=True, color="FFFFFF", size=11),
        "header_fill": PatternFill("solid", fgColor="1F4E78"),
        "subheader_fill": PatternFill("solid", fgColor="4472C4"),
        "blue_font": Font(name="Arial", color="1F4E78", size=10),
        "black_font": Font(name="Arial", color="1F4E78", size=10),
        "green_font": Font(name="Arial", color="1F4E78", size=10),
        "red_font": Font(name="Arial", color="1F4E78", size=10),
        "orange_font": Font(name="Arial", color="1F4E78", size=10),
        "gray_font": Font(name="Arial", color="1F4E78", size=10),
        "center_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left_align": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "thin_border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        ),
    }


def _set_header(ws, row, values, styles, fill_key="header_fill", start_col=1):
    for c, v in enumerate(values, start_col):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = styles["header_font"]
        cell.fill = styles[fill_key]
        cell.alignment = styles["center_align"]
        cell.border = styles["thin_border"]


def _write_row(ws, row, values, styles, is_input=False, is_warn=False, is_note=False, align=None):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        if is_input:
            cell.font = styles["blue_font"]
        elif is_warn:
            cell.font = styles["orange_font"]
        elif is_note:
            cell.font = styles["gray_font"]
        else:
            cell.font = styles["black_font"]
        cell.alignment = align or (styles["left_align"] if isinstance(v, str) else styles["center_align"])
        cell.border = styles["thin_border"]


def _set_column_widths(ws, widths):
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _autofit_columns(ws, min_width=10, max_width=40):
    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted_width


def _parse_date(date_str):
    if not date_str:
        return datetime.now().strftime("%Y%m%d")
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").strftime("%Y%m%d")
    except ValueError:
        return date_str


# ---------------------------------------------------------------------------
# Preset loading
# ---------------------------------------------------------------------------
def load_presets():
    path = DATA_DIR / "sales_weekly_templates.json"
    if not path.exists():
        return {}
    return load_json(path)


def get_preset(name):
    presets = load_presets()
    for key, value in presets.items():
        if key == name or value.get("team_type") == name or value.get("team_name") == name:
            return dict(value)
    return None


def list_presets():
    return list(load_presets().keys())


def build_params(args):
    presets = load_presets()
    preset_name = args.preset or "通用销售团队"
    preset = None
    if preset_name:
        preset = get_preset(preset_name)
    if not preset:
        preset = dict(presets.get("通用销售团队", {}))

    params = {
        "team_type": preset.get("team_type", "通用销售团队"),
        "team_name": args.team or preset.get("team_name", "销售部"),
        "week": args.week or _current_week_label(),
        "sales_target": args.sales_target or preset.get("default_sales_target", 600000),
        "avg_deal_value": preset.get("avg_deal_value", 5000),
        "sales_cycle_days": preset.get("sales_cycle_days", 14),
        "core_kpis": preset.get("core_kpis", []),
        "kpi_weights": preset.get("kpi_weights", {}),
        "pipeline_stages": preset.get("pipeline_stages", []),
        "issue_categories": preset.get("issue_categories", []),
        "plan_priorities": preset.get("plan_priorities", []),
        "author": args.author or "",
        "date": args.date or datetime.now().strftime("%Y-%m-%d"),
        "key_conclusions": "",
    }
    return params


def _current_week_label():
    today = datetime.now()
    year, week, _ = today.isocalendar()
    return f"{year}年第{week}周"


# ---------------------------------------------------------------------------
# Workbook building
# ---------------------------------------------------------------------------
def build_workbook(params, output_path):
    wb = Workbook()
    styles = _base_styles()

    build_cover(wb, params, styles)
    build_performance(wb, params, styles)
    build_pipeline(wb, params, styles)
    build_issues(wb, params, styles)
    build_next_week_plan(wb, params, styles)

    wb.save(output_path)
    return output_path


def build_cover(wb, params, styles):
    ws = wb.active
    ws.title = "封面汇总"

    team_name = params["team_name"]
    week = params["week"]
    author = params["author"]
    date_str = params["date"]

    ws["B2"] = "销售团队标准周报"
    ws["B2"].font = styles["title_font"]
    ws.merge_cells("B2:F2")
    ws["B2"].alignment = styles["left_align"]

    info = [
        ["团队名称", team_name],
        ["报表周期", week],
        ["填写人", author or "待填写"],
        ["填写日期", date_str],
    ]
    for i, (label, value) in enumerate(info, 4):
        ws.cell(row=i, column=2, value=label).font = styles["section_font"]
        ws.cell(row=i, column=2).alignment = styles["left_align"]
        ws.cell(row=i, column=3, value=value).font = styles["blue_font"]
        ws.cell(row=i, column=3).alignment = styles["left_align"]

    ws["B9"] = "本周关键结论"
    ws["B9"].font = styles["section_font"]
    ws.merge_cells("B9:F9")

    conclusions = [
        "1. 本周业绩达成情况：",
        "2. 重点商机推进亮点：",
        "3. 主要风险与问题：",
        "4. 下周重点关注：",
    ]
    for i, text in enumerate(conclusions, 10):
        ws.cell(row=i, column=2, value=text).font = styles["black_font"]
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        ws.cell(row=i, column=2).alignment = styles["left_align"]

    ws["B16"] = "说明"
    ws["B16"].font = styles["section_font"]
    notes = [
        "• 本模板由 prajna 企智 prajna-sales-weekly-report 技能自动生成。",
        "• 蓝色字体单元格为建议填写项，黑色为说明或计算结果。",
        "• 核心业绩表已内置公式，请直接录入目标值与实际值。",
    ]
    for i, note in enumerate(notes, 17):
        ws.cell(row=i, column=2, value=note).font = styles["gray_font"]
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        ws.cell(row=i, column=2).alignment = styles["left_align"]

    _set_column_widths(ws, [4, 18, 30, 18, 18, 18])


def build_performance(wb, params, styles):
    ws = wb.create_sheet("核心业绩数据")

    # Section 1: overall target / actual / completion
    ws["A1"] = "一、本周业绩总览"
    ws["A1"].font = styles["section_font"]
    ws.merge_cells("A1:H1")

    headers1 = ["指标", "本周目标", "本周实际", "完成率", "上周同期", "环比", "上月同期", "同比"]
    _set_header(ws, 2, headers1, styles)
    target = params["sales_target"]
    actual = round(target * 0.88, 2)
    last_week = round(target * 0.82, 2)
    last_month = round(target * 0.85, 2)
    rows1 = [
        ["销售额", target, actual, "=C3/B3", last_week, "=(C3-D3)/D3", last_month, "=(C3-H3)/H3"],
        ["回款额", round(target * 0.75, 2), round(target * 0.68, 2), "=C4/B4", round(target * 0.70, 2), "=(C4-D4)/D4", round(target * 0.72, 2), "=(C4-H4)/H4"],
        ["签约客户数", max(1, int(target / params.get("avg_deal_value", 5000))), max(1, int(target / params.get("avg_deal_value", 5000) * 0.9)), "=C5/B5", "-", "-", "-", "-"],
    ]
    for i, row in enumerate(rows1, 3):
        _write_row(ws, i, row, styles, is_input=True)
    ws.freeze_panes = "A3"

    # Format percentages
    for r in range(3, 6):
        ws.cell(row=r, column=4).number_format = "0.00%"
        ws.cell(row=r, column=6).number_format = "0.00%"
        ws.cell(row=r, column=8).number_format = "0.00%"

    # Section 2: ranking
    ws["A8"] = "二、业绩排名（Top 5）"
    ws["A8"].font = styles["section_font"]
    ws.merge_cells("A8:F8")

    headers2 = ["排名", "销售人员", "本周业绩", "目标", "完成率", "备注"]
    _set_header(ws, 9, headers2, styles)
    sample_names = ["张三", "李四", "王五", "赵六", "孙七"]
    for i, name in enumerate(sample_names, 10):
        person_target = round(target / 5, 2)
        person_actual = round(person_target * (1.05 - i * 0.08), 2)
        row = [i, name, person_actual, person_target, f"=C{i+1}/D{i+1}", ""]
        _write_row(ws, i, row, styles, is_input=(name == "张三"))
        ws.cell(row=i, column=5).number_format = "0.00%"

    # Section 3: core KPI details
    start_row = 17
    ws.cell(row=start_row, column=1, value="三、核心 KPI 明细").font = styles["section_font"]
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)

    headers3 = ["序号", "KPI 指标", "单位", "权重", "目标值", "实际值", "达成率", "得分", "备注"]
    _set_header(ws, start_row + 1, headers3, styles)

    core_kpis = params.get("core_kpis", [])
    if not core_kpis:
        core_kpis = [
            {"name": "销售额", "unit": "元", "target": params["sales_target"], "weight": 0.4},
            {"name": "回款额", "unit": "元", "target": round(params["sales_target"] * 0.75, 2), "weight": 0.25},
            {"name": "新客户数", "unit": "家", "target": 20, "weight": 0.15},
            {"name": "客户拜访量", "unit": "次", "target": 60, "weight": 0.1},
            {"name": "商机推进", "unit": "个", "target": 15, "weight": 0.1},
        ]

    for i, kpi in enumerate(core_kpis, start_row + 2):
        target_val = kpi.get("target", 0)
        weight = kpi.get("weight", 0)
        actual_val = round(target_val * (0.85 + (i % 3) * 0.05), 2)
        row_idx = i
        row = [
            i - start_row - 1,
            kpi.get("name", ""),
            kpi.get("unit", ""),
            weight,
            target_val,
            actual_val,
            f"=F{row_idx + 1}/E{row_idx + 1}",
            f"=G{row_idx + 1}*D{row_idx + 1}*100",
            "",
        ]
        _write_row(ws, i, row, styles, is_input=True)
        ws.cell(row=i, column=4).number_format = "0.00%"
        ws.cell(row=i, column=7).number_format = "0.00%"
        ws.cell(row=i, column=8).number_format = "0.00"

    total_row = start_row + 2 + len(core_kpis)
    ws.cell(row=total_row, column=1, value="合计").font = styles["black_font"]
    for c in range(2, 9):
        ws.cell(row=total_row, column=c, value="").border = styles["thin_border"]
    ws.cell(row=total_row, column=8, value=f"=SUM(H{start_row + 2}:H{total_row - 1})").font = styles["black_font"]
    ws.cell(row=total_row, column=8).number_format = "0.00"
    for c in range(1, 10):
        ws.cell(row=total_row, column=c).border = styles["thin_border"]
        ws.cell(row=total_row, column=c).font = Font(bold=True, name="Arial", size=10)

    _autofit_columns(ws, min_width=10, max_width=22)
    ws.column_dimensions["B"].width = 22


def build_pipeline(wb, params, styles):
    ws = wb.create_sheet("重点商机进展")

    ws["A1"] = "重点商机最新进展"
    ws["A1"].font = styles["section_font"]
    ws.merge_cells("A1:I1")

    headers = [
        "序号", "客户名称", "行业", "阶段", "预计金额(元)", "预计成交时间",
        "负责人", "本周进展", "下一步动作", "风险/阻碍"
    ]
    _set_header(ws, 2, headers, styles)
    ws.freeze_panes = "A3"

    stages = params.get("pipeline_stages", ["线索", "接触", "需求确认", "报价", "谈判", "成交"])
    sample_companies = [
        ("华智科技", "互联网", stages[2] if len(stages) > 2 else "方案演示", 120000, "2026-08-15", "张三"),
        ("锐行零售", "零售", stages[3] if len(stages) > 3 else "商务谈判", 80000, "2026-08-20", "李四"),
        ("远景制造", "制造业", stages[1] if len(stages) > 1 else "需求沟通", 200000, "2026-09-01", "王五"),
        ("云帆 SaaS", "企业服务", stages[4] if len(stages) > 4 else "合同签署", 150000, "2026-08-10", "赵六"),
    ]
    for i, (company, industry, stage, amount, close_date, owner) in enumerate(sample_companies, 3):
        row = [i - 2, company, industry, stage, amount, close_date, owner, "", "", ""]
        _write_row(ws, i, row, styles, is_input=(i == 3))
        ws.cell(row=i, column=5).number_format = '#,##0'

    _autofit_columns(ws, min_width=10, max_width=26)
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["H"].width = 26
    ws.column_dimensions["I"].width = 26
    ws.column_dimensions["J"].width = 24


def build_issues(wb, params, styles):
    ws = wb.create_sheet("问题分析")

    ws["A1"] = "本周遇到的问题分析"
    ws["A1"].font = styles["section_font"]
    ws.merge_cells("A1:G1")

    headers = ["序号", "问题描述", "影响范围", "根因分析", "解决方案", "负责人", "截止日期", "状态"]
    _set_header(ws, 2, headers, styles)
    ws.freeze_panes = "A3"

    issue_categories = params.get("issue_categories", [])
    sample_issues = [
        (issue_categories[0] if len(issue_categories) > 0 else "市场竞争加剧", "整体签单周期", "竞品降价", "强化价值销售话术", "张三", "2026-07-25", "进行中"),
        (issue_categories[1] if len(issue_categories) > 1 else "客户需求变更", "重点项目 A", "需求边界不清", "召开需求确认会", "李四", "2026-07-22", "待跟进"),
        (issue_categories[2] if len(issue_categories) > 2 else "内部交付延迟", "客户交付", "资源不足", "协调产研资源", "王五", "2026-07-30", "待启动"),
    ]
    statuses = ["待启动", "进行中", "待验证", "已完成", "已关闭"]
    for i, (desc, impact, root, solution, owner, deadline, status) in enumerate(sample_issues, 3):
        row = [i - 2, desc, impact, root, solution, owner, deadline, status]
        _write_row(ws, i, row, styles, is_input=(i == 3), is_warn=(status in ("待启动", "待跟进")))

    # Data validation hint for status column
    ws["H2"].comment = None

    _autofit_columns(ws, min_width=10, max_width=28)
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 26


def build_next_week_plan(wb, params, styles):
    ws = wb.create_sheet("下周工作计划")

    ws["A1"] = "下周工作重点计划"
    ws["A1"].font = styles["section_font"]
    ws.merge_cells("A1:H1")

    headers = [
        "序号", "工作事项", "优先级", "目标/指标", "负责人",
        "协作方", "截止日期", "所需资源", "完成标准"
    ]
    _set_header(ws, 2, headers, styles)
    ws.freeze_panes = "A3"

    priorities = ["P0", "P1", "P2"]
    plan_priorities = params.get("plan_priorities", [])
    sample_plans = [
        (plan_priorities[0] if len(plan_priorities) > 0 else "重点客户攻坚", "P0", "签约 3 家头部客户", "张三", "售前/产研", "2026-07-31", "方案支持", "合同签署"),
        (plan_priorities[1] if len(plan_priorities) > 1 else "新行业拓展", "P1", "完成 2 场行业沙龙", "李四", "市场", "2026-08-05", "市场预算", "获取 50 条线索"),
        (plan_priorities[2] if len(plan_priorities) > 2 else "销售培训", "P1", "人均培训 4 小时", "王五", "HR", "2026-08-02", "培训场地", "培训覆盖率 100%"),
        (plan_priorities[3] if len(plan_priorities) > 3 else "流程优化", "P2", "CRM 字段补齐率 95%", "赵六", "运营", "2026-08-10", "CRM 权限", "周会通报"),
    ]
    for i, (item, priority, target, owner, collaborators, deadline, resources, criteria) in enumerate(sample_plans, 3):
        row = [i - 2, item, priority, target, owner, collaborators, deadline, resources, criteria]
        is_p0 = priority == "P0"
        _write_row(ws, i, row, styles, is_input=True, is_warn=is_p0)

    _autofit_columns(ws, min_width=10, max_width=26)
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["I"].width = 24


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="生成销售团队标准周报 Excel 模板"
    )
    parser.add_argument(
        "--preset", "-p",
        help=f"团队预设类型，可选：{', '.join(list_presets())}，默认通用销售团队"
    )
    parser.add_argument(
        "--team", "-t",
        help="团队名称（覆盖预设中的默认值）"
    )
    parser.add_argument(
        "--week", "-w",
        help="报表周期，例如 '2026年第30周'，默认当前周"
    )
    parser.add_argument(
        "--sales-target", "-s", type=float,
        help="本周销售目标金额，覆盖预设默认值"
    )
    parser.add_argument(
        "--author", "-a",
        help="填写人姓名"
    )
    parser.add_argument(
        "--date", "-d",
        help="填写日期，格式 YYYY-MM-DD，默认今天"
    )
    parser.add_argument(
        "--output", "-o",
        help="输出文件路径（覆盖默认路径）"
    )
    parser.add_argument(
        "--list-presets", action="store_true",
        help="列出所有可用预设"
    )
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    if args.list_presets:
        print("可用团队预设：")
        for name in list_presets():
            print(f"  - {name}")
        return 0

    params = build_params(args)

    if args.output:
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
    else:
        HISTORY_DIR.mkdir(parents=True, exist_ok=True)
        date_suffix = _parse_date(params["date"])
        safe_team = _safe_filename(params["team_name"])
        filename = f"prajna_销售周报_{safe_team}_{date_suffix}.xlsx"
        output_path = HISTORY_DIR / filename

    build_workbook(params, output_path)
    print(f"已生成销售周报：{output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
