#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
prajna 企智 - 服装厂小组长岗位职责与 KPI 考核表生成器

生成一份 Excel 工作簿，包含：
1. 使用说明
2. 岗位职责矩阵（上游部门、下游部门、岗位职责、考核指标、量化标准）
3. KPI 考核指标（8 项核心 KPI）
4. 绩效评分记录表（支持自动计算加权得分）

依赖：openpyxl（未安装时脚本会提示安装命令）
"""

import argparse
import os
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    print(
        "[错误] 缺少 openpyxl，请执行：pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)


SKILL_NAME = "prajna-clothing-teamleader-duty"
DEFAULT_OUTPUT_DIR = Path.home() / ".prajna" / SKILL_NAME / "samples"

THEME_BLUE = "1F4E78"
THEME_LIGHT_BLUE = "D9E1F2"
THEME_BG = "F2F2F2"

HEADER_FILL = PatternFill(start_color=THEME_BLUE, end_color=THEME_BLUE, fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color=THEME_LIGHT_BLUE, end_color=THEME_LIGHT_BLUE, fill_type="solid")
NORMAL_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B7B7B7"),
    right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"),
    bottom=Side(style="thin", color="B7B7B7"),
)


PRESETS = {
    "服装缝纫组": {"team": "缝纫一组", "factory": "成衣A厂"},
    "服装裁剪组": {"team": "裁剪组", "factory": "成衣A厂"},
    "服装包装组": {"team": "包装组", "factory": "成衣A厂"},
    "通用": {"team": "车间小组", "factory": "XX服装厂"},
}


RESPONSIBILITIES = [
    {
        "module": "生产执行",
        "upstream": "生产计划部/PMC",
        "downstream": "后道质检/成品仓",
        "duty": "根据日/周生产计划，合理分配工序与人员，组织本组保质保量完成裁剪/缝纫/包装任务，并按时交付下道工序。",
        "indicator": "产量达成率",
        "standard": "≥95%（实际产量/计划产量）",
    },
    {
        "module": "生产执行",
        "upstream": "物料仓/采购部",
        "downstream": "后道质检",
        "duty": "开产前确认物料齐套情况，发现缺料、错料、色差等问题及时上报并跟进解决，避免停线待料。",
        "indicator": "物料齐套率",
        "standard": "≥98%（当日齐套批次/总批次）",
    },
    {
        "module": "生产执行",
        "upstream": "技术/工艺部",
        "downstream": "后道质检",
        "duty": "监督员工按标准工时、工序顺序操作，控制生产节拍，减少瓶颈工序堆积，提升人均效率。",
        "indicator": "人均小时产量",
        "standard": "≥定额标准的95%",
    },
    {
        "module": "质量管控",
        "upstream": "技术/工艺部/品控部",
        "downstream": "后道质检/客户",
        "duty": "执行首件检验、巡检及成品自检制度，确保产品符合工艺单、样板及客户质量要求。",
        "indicator": "质量合格率",
        "standard": "≥98%（合格数/总检验数）",
    },
    {
        "module": "质量管控",
        "upstream": "后道质检",
        "downstream": "返修组/生产计划",
        "duty": "对不合格品进行标识、隔离、记录，组织本组返修并跟踪返修结果，防止批量性质量问题发生。",
        "indicator": "返修率",
        "standard": "≤2%（返修件数/总产出件数）",
    },
    {
        "module": "质量管控",
        "upstream": "品控部/客户",
        "downstream": "技术/工艺部",
        "duty": "收集、反馈质量异常，参与原因分析并落实纠正预防措施（CAPA）。",
        "indicator": "批量质量问题",
        "standard": "0 起/月（同一批次≥10件同类不合格）",
    },
    {
        "module": "人员管理",
        "upstream": "人力资源部",
        "downstream": "生产计划部/车间主任",
        "duty": "负责本组排班、考勤、绩效面谈及日常纪律管理，确保生产人力满足排产需求。",
        "indicator": "人员出勤率",
        "standard": "≥96%（实际出勤工时/应出勤工时）",
    },
    {
        "module": "人员管理",
        "upstream": "人力资源部/培训部",
        "downstream": "生产计划部",
        "duty": "对新进员工进行岗位技能培训、安全交底及师带徒跟踪，帮助其快速达到上岗标准。",
        "indicator": "新员工培训合格率",
        "standard": "≥90%（培训合格人数/培训总人数）",
    },
    {
        "module": "人员管理",
        "upstream": "人力资源部",
        "downstream": "车间主任",
        "duty": "关注员工思想动态与留任意愿，及时化解矛盾，配合 HR 降低非自愿流失。",
        "indicator": "员工流失率",
        "standard": "≤5%/月（离职人数/月初在岗人数）",
    },
    {
        "module": "设备与安全",
        "upstream": "设备部",
        "downstream": "生产计划部/安全员",
        "duty": "组织本组设备日常点检、保养与简单故障排除，及时报修重大故障并记录停机原因。",
        "indicator": "设备故障停机时间",
        "standard": "≤4小时/月（本组责任停机）",
    },
    {
        "module": "设备与安全",
        "upstream": "安全/行政部",
        "downstream": "生产计划部",
        "duty": "开展班前安全提醒，监督劳保用品佩戴及安全操作规程执行，排查现场安全隐患。",
        "indicator": "安全事故",
        "standard": "0 起/月（含轻伤及以上及火警）",
    },
    {
        "module": "成本与现场",
        "upstream": "采购部/物料仓",
        "downstream": "财务部/生产计划",
        "duty": "控制面辅料、线、标签等物料损耗，余料及时退仓，降低单件物料成本。",
        "indicator": "物料损耗率",
        "standard": "≤3%（实际损耗/标准用量）",
    },
    {
        "module": "成本与现场",
        "upstream": "行政/车间主任",
        "downstream": "生产计划部",
        "duty": "推行 5S 管理，保持工位、通道、裁片/半成品定置定位，营造有序生产环境。",
        "indicator": "5S检查得分",
        "standard": "≥85分/次（车间5S评比）",
    },
    {
        "module": "成本与现场",
        "upstream": "生产计划部",
        "downstream": "财务部/HR",
        "duty": "准确记录本组工时、产量及异常原因，为成本核算、计件工资提供可靠数据。",
        "indicator": "工时填报准确率",
        "standard": "≥98%（准确填报条数/总填报条数）",
    },
]


KPI_DATA = [
    {
        "id": "KPI-01",
        "name": "产量达成率",
        "weight": 25,
        "target": "≥95%",
        "formula": "实际产量/计划产量×100%",
        "source": "MES/生产日报",
        "rule": "每低1%扣2分，低于85%得0分",
        "cycle": "月度",
    },
    {
        "id": "KPI-02",
        "name": "质量合格率",
        "weight": 25,
        "target": "≥98%",
        "formula": "合格数/总检验数×100%",
        "source": "品检日报/返修记录",
        "rule": "每低0.5%扣3分，低于93%得0分",
        "cycle": "月度",
    },
    {
        "id": "KPI-03",
        "name": "返修率",
        "weight": 10,
        "target": "≤2%",
        "formula": "返修件数/总产出件数×100%",
        "source": "后道质检/返修记录",
        "rule": "每超0.5%扣3分，高于5%得0分",
        "cycle": "月度",
    },
    {
        "id": "KPI-04",
        "name": "人员出勤率",
        "weight": 10,
        "target": "≥96%",
        "formula": "实际出勤工时/应出勤工时×100%",
        "source": "考勤系统/排班表",
        "rule": "每低1%扣2分，低于90%得0分",
        "cycle": "月度",
    },
    {
        "id": "KPI-05",
        "name": "员工流失率",
        "weight": 5,
        "target": "≤5%/月",
        "formula": "离职人数/月初在岗人数×100%",
        "source": "HR系统/离职审批",
        "rule": "每超1%扣2分，高于10%得0分",
        "cycle": "月度",
    },
    {
        "id": "KPI-06",
        "name": "设备故障停机时间",
        "weight": 10,
        "target": "≤4小时/月",
        "formula": "∑本组责任停机小时数",
        "source": "设备维修单/MES",
        "rule": "每超1小时扣2分，高于12小时得0分",
        "cycle": "月度",
    },
    {
        "id": "KPI-07",
        "name": "安全事故",
        "weight": 10,
        "target": "0 起/月",
        "formula": "本组责任安全事故起数",
        "source": "安全报告/事故记录",
        "rule": "每发生一起扣5分，重大事故0分",
        "cycle": "月度",
    },
    {
        "id": "KPI-08",
        "name": "物料损耗率",
        "weight": 5,
        "target": "≤3%",
        "formula": "实际损耗/标准用量×100%",
        "source": "领料单/退料单/成本核算",
        "rule": "每超0.5%扣2分，高于6%得0分",
        "cycle": "月度",
    },
]


MODULE_COLORS = {
    "生产执行": "E2EFDA",
    "质量管控": "FCE4D6",
    "人员管理": "DDEBF7",
    "设备与安全": "FFF2CC",
    "成本与现场": "E1D5E7",
}


def _safe_filename(text: str) -> str:
    """移除或替换文件名中的非法字符。"""
    return re.sub(r'[\\/:*?"<>|]+', "_", text).strip()


def _apply_header_row(ws, row_idx, headers, widths=None):
    """给指定行写入表头并设置样式。"""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    if widths:
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width


def _style_data_cell(cell, font=None, align=None, fill=None):
    cell.font = font or NORMAL_FONT
    cell.alignment = align or LEFT_ALIGN
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill


def _add_readme_sheet(wb, metadata):
    ws = wb.active
    ws.title = "使用说明"
    ws.append([])
    ws.append(["服装厂小组长岗位职责与 KPI 考核表"])
    ws.append([])
    ws.append(["工厂名称", metadata["factory"]])
    ws.append(["班组名称", metadata["team"]])
    ws.append(["考核周期", metadata["month"]])
    ws.append(["编制人", metadata["author"]])
    ws.append(["生成日期", metadata["date"]])
    ws.append([])
    ws.append(["本 Excel 包含三张工作表："])
    ws.append(["1. 岗位职责矩阵", "按生产执行、质量管控、人员管理、设备与安全、成本与现场五大模块列出小组长岗位职责及上下游协作关系。"])
    ws.append(["2. KPI 考核指标", "8 项核心 KPI 的权重、目标值、计算公式、数据来源、评分规则与考核周期。"])
    ws.append(["3. 绩效评分记录", "在「实际完成」列填写实际值，在「单项得分」列人工评分，「加权得分」会自动计算。"])
    ws.append([])
    ws.append(["填写须知"])
    ws.append(["1.", "量化标准中的目标值为行业参考值，企业应根据自身产能、工艺、客户要求进行调整。"])
    ws.append(["2.", "岗位职责可根据班组实际分工进行增删，上游/下游部门名称以企业内部组织架构为准。"])
    ws.append(["3.", "绩效评分记录表中的「单项得分」为百分制，需按 KPI 考核指标中的评分规则人工评定。"])
    ws.append([])
    ws.append(["免责声明"])
    ws.append(["本模板由 prajna 企智人工智能生成，仅供业务参考，不构成正式绩效考核或劳动用工依据，使用前请由企业相关部门审核确认。"])

    # 合并标题行
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    title_cell = ws.cell(row=2, column=1)
    title_cell.font = Font(name="微软雅黑", size=16, bold=True, color=THEME_BLUE)
    title_cell.alignment = CENTER_ALIGN

    for row in ws.iter_rows(min_row=4, max_row=8, min_col=1, max_col=2):
        row[0].font = BOLD_FONT
        row[0].alignment = LEFT_ALIGN
        row[1].alignment = LEFT_ALIGN

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90


def _add_responsibility_sheet(wb):
    ws = wb.create_sheet("岗位职责矩阵")
    headers = ["模块", "上游部门", "下游部门", "岗位职责", "考核指标", "量化标准"]
    widths = [14, 18, 18, 60, 16, 24]
    _apply_header_row(ws, 1, headers, widths)
    ws.freeze_panes = "A2"

    for row_idx, item in enumerate(RESPONSIBILITIES, 2):
        fill = PatternFill(
            start_color=MODULE_COLORS.get(item["module"], "FFFFFF"),
            end_color=MODULE_COLORS.get(item["module"], "FFFFFF"),
            fill_type="solid",
        )
        values = [
            item["module"],
            item["upstream"],
            item["downstream"],
            item["duty"],
            item["indicator"],
            item["standard"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            align = CENTER_ALIGN if col_idx in (1, 2, 3, 5, 6) else LEFT_ALIGN
            _style_data_cell(cell, align=align, fill=fill)
        ws.row_dimensions[row_idx].height = 48


def _add_kpi_sheet(wb):
    ws = wb.create_sheet("KPI考核指标")
    headers = ["KPI编号", "KPI名称", "权重(%)", "目标值", "计算方式", "数据来源", "评分规则", "考核周期"]
    widths = [10, 14, 10, 14, 26, 18, 32, 10]
    _apply_header_row(ws, 1, headers, widths)
    ws.freeze_panes = "A2"

    total_weight = 0
    for row_idx, item in enumerate(KPI_DATA, 2):
        total_weight += item["weight"]
        values = [
            item["id"],
            item["name"],
            item["weight"],
            item["target"],
            item["formula"],
            item["source"],
            item["rule"],
            item["cycle"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            align = CENTER_ALIGN if col_idx in (1, 3, 4, 8) else LEFT_ALIGN
            _style_data_cell(cell, align=align)
        ws.row_dimensions[row_idx].height = 36

    # 权重合计校验行
    total_row = len(KPI_DATA) + 2
    ws.cell(row=total_row, column=1, value="合计")
    ws.cell(row=total_row, column=1).font = BOLD_FONT
    ws.cell(row=total_row, column=1).alignment = CENTER_ALIGN
    ws.cell(row=total_row, column=1).border = THIN_BORDER

    ws.cell(row=total_row, column=2, value="权重合计应=100%")
    ws.cell(row=total_row, column=2).font = BOLD_FONT
    ws.cell(row=total_row, column=2).alignment = CENTER_ALIGN
    ws.cell(row=total_row, column=2).border = THIN_BORDER

    weight_cell = ws.cell(row=total_row, column=3, value=total_weight)
    weight_cell.font = BOLD_FONT
    weight_cell.alignment = CENTER_ALIGN
    weight_cell.border = THIN_BORDER
    weight_cell.fill = PatternFill(
        start_color=THEME_LIGHT_BLUE, end_color=THEME_LIGHT_BLUE, fill_type="solid"
    )

    for col in range(4, 9):
        cell = ws.cell(row=total_row, column=col, value="")
        cell.border = THIN_BORDER


def _add_score_sheet(wb):
    ws = wb.create_sheet("绩效评分记录")
    headers = ["KPI编号", "KPI名称", "权重(%)", "目标值", "实际完成", "单项得分", "加权得分", "备注"]
    widths = [10, 14, 10, 14, 14, 12, 12, 24]
    _apply_header_row(ws, 1, headers, widths)
    ws.freeze_panes = "A2"

    for row_idx, item in enumerate(KPI_DATA, 2):
        values = [item["id"], item["name"], item["weight"], item["target"], "", "", "", ""]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            align = CENTER_ALIGN if col_idx in (1, 3, 4, 5, 6, 7) else LEFT_ALIGN
            _style_data_cell(cell, align=align)

        # 加权得分公式：=C2*F2/100，当权重与单项得分均非空时计算
        weight_col = 3
        score_col = 6
        weighted_col = 7
        formula = (
            f'=IF(AND({get_column_letter(weight_col)}{row_idx}<>"",'
            f'{get_column_letter(score_col)}{row_idx}<>""),'
            f'ROUND({get_column_letter(weight_col)}{row_idx}*{get_column_letter(score_col)}{row_idx}/100,2),"")'
        )
        ws.cell(row=row_idx, column=weighted_col, value=formula)

    # 汇总行
    total_row = len(KPI_DATA) + 2
    ws.cell(row=total_row, column=1, value="合计")
    ws.cell(row=total_row, column=1).font = BOLD_FONT
    ws.cell(row=total_row, column=1).alignment = CENTER_ALIGN
    ws.cell(row=total_row, column=1).border = THIN_BORDER

    ws.cell(row=total_row, column=2, value="绩效总得分")
    ws.cell(row=total_row, column=2).font = BOLD_FONT
    ws.cell(row=total_row, column=2).alignment = CENTER_ALIGN
    ws.cell(row=total_row, column=2).border = THIN_BORDER

    for col in range(3, 6):
        ws.cell(row=total_row, column=col, value="").border = THIN_BORDER

    total_formula = f'=IF(COUNTA(G2:G{total_row - 1})>0,SUM(G2:G{total_row - 1}),"")'
    total_cell = ws.cell(row=total_row, column=7, value=total_formula)
    total_cell.font = BOLD_FONT
    total_cell.alignment = CENTER_ALIGN
    total_cell.border = THIN_BORDER
    total_cell.fill = PatternFill(
        start_color=THEME_LIGHT_BLUE, end_color=THEME_LIGHT_BLUE, fill_type="solid"
    )
    ws.cell(row=total_row, column=8, value="").border = THIN_BORDER


def generate_excel(output_path, metadata):
    wb = Workbook()
    _add_readme_sheet(wb, metadata)
    _add_responsibility_sheet(wb)
    _add_kpi_sheet(wb)
    _add_score_sheet(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def build_metadata(args):
    preset = PRESETS.get(args.preset, PRESETS["服装缝纫组"])
    team = args.team or preset["team"]
    factory = args.factory or preset["factory"]
    month = args.month or datetime.now().strftime("%Y-%m")
    author = args.author or os.environ.get("USER", "prajna")
    date_str = datetime.now().strftime("%Y-%m-%d")
    return {
        "preset": args.preset,
        "team": team,
        "factory": factory,
        "month": month,
        "author": author,
        "date": date_str,
    }


def main():
    parser = argparse.ArgumentParser(
        description="生成服装厂小组长岗位职责与 KPI 考核 Excel 模板"
    )
    parser.add_argument(
        "--preset",
        "-p",
        default="服装缝纫组",
        help=f"班组预设，可选：{', '.join(PRESETS.keys())}",
    )
    parser.add_argument("--team", "-t", help="班组名称，覆盖预设默认值")
    parser.add_argument("--factory", "-f", help="工厂名称，默认成衣A厂")
    parser.add_argument("--month", "-m", help="考核周期，如 2026-07，默认当前月份")
    parser.add_argument("--author", "-a", help="编制人姓名")
    parser.add_argument("--output", "-o", help="输出文件完整路径，默认保存到 ~/.prajna/prajna-clothing-teamleader-duty/samples/")
    parser.add_argument("--list-presets", action="store_true", help="列出所有可用预设")

    args = parser.parse_args()

    if args.list_presets:
        print("可用预设：")
        for name, info in PRESETS.items():
            print(f"  - {name}: 默认班组={info['team']}, 默认工厂={info['factory']}")
        return 0

    metadata = build_metadata(args)

    if args.output:
        output_path = Path(args.output).expanduser().resolve()
    else:
        factory_part = _safe_filename(metadata["factory"])
        team_part = _safe_filename(metadata["team"])
        month_part = _safe_filename(metadata["month"])
        filename = f"prajna_服装厂小组长岗位职责与KPI_{factory_part}_{team_part}_{month_part}.xlsx"
        output_path = DEFAULT_OUTPUT_DIR / filename

    generate_excel(output_path, metadata)
    print(f"已生成：{output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
